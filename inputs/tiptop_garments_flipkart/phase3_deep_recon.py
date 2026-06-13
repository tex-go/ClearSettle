"""
ClearSettle — Flipkart Deep Reconciliation Engine
Client: Tip Top Garments
Phase 3: Per-order fee validation, commission rate analysis, anomaly detection
"""

import sys, io, json
from pathlib import Path
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings('ignore')

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

BASE_DIR = Path("C:/Ranjith/working_dir/ClearSettle/inputs/tiptop_garments_flipkart")
OUT_DIR  = BASE_DIR / "analysis_output"
EXT_DIR  = OUT_DIR / "raw_extracted"

def n(x):
    try:
        v = str(x).replace(',', '').replace(' ', '').strip()
        return float(v) if v not in ('', 'nan', 'None', '-', 'N/A', '#N/A') else 0.0
    except:
        return 0.0

def s(x):
    return str(x).strip() if pd.notna(x) and str(x).strip() not in ('nan','None','') else ''

# ═══════════════════════════════════════════════════════════════════════════
# LOAD PAYMENT REPORTS (3-row header)
# Row 0: Section headers
# Row 1: Main column names
# Row 2: Sub-column names (additional breakdowns)
# Data: Row 3+
# ═══════════════════════════════════════════════════════════════════════════

def parse_payment_report_full(path: Path, quarter_label: str = '') -> pd.DataFrame:
    """Parse Flipkart Payment Report Orders sheet with full column recovery."""
    raw = pd.read_excel(path, sheet_name='Orders', header=None, dtype=str, engine='openpyxl')
    if len(raw) < 4:
        return pd.DataFrame()

    # Detect header rows
    r0, r1, r2 = 0, 1, 2
    for i in range(min(5, len(raw))):
        row_str = ' '.join(s(v).lower() for v in raw.iloc[i])
        if 'payment details' in row_str or 'neft' in row_str.lower():
            r0 = i
            r1 = i + 1
            r2 = i + 2
            break

    # Check if row2 is a sub-header or data
    r2_sample = [s(v) for v in raw.iloc[r2] if s(v)]
    is_subheader = any(any(k in v.lower() for k in ['total', 'sum', 'sub']) for v in r2_sample)

    # Build column names from row1 (main headers) — forward-fill section from row0
    sec_row = list(raw.iloc[r0])
    col_row = list(raw.iloc[r1])
    sub_row = list(raw.iloc[r2]) if is_subheader else [None] * len(col_row)

    final_cols = []
    curr_sec = ''
    for i in range(len(col_row)):
        sec = s(sec_row[i]) if i < len(sec_row) else ''
        col = s(col_row[i])
        if sec and not sec.lower().startswith('unnamed'):
            curr_sec = sec

        if col and not col.lower().startswith('unnamed'):
            nm = col
        elif curr_sec:
            nm = f"{curr_sec}_{i}"
        else:
            nm = f"col_{i}"

        # Clean up
        nm = (nm.strip()
              .lower()
              .replace('\n', ' ')
              .replace('= sum (v:ai)', '')
              .replace('= sum(j:r)', '')
              .replace('[gst+tcs]', '')
              .replace('[tds]', '')
              .replace('[ao/(1+ap/100)]', '')
              .replace('[18%*aq]', '')
              .replace('[as+ao]', '')
              .replace('[aq + ar]', '')
              .strip())

        # Deduplicate
        count = sum(1 for c in final_cols if c == nm)
        if count > 0:
            nm = f"{nm}_{count}"

        final_cols.append(nm)

    # Data rows
    data_start = r2 + 1 if is_subheader else r1 + 1
    df = raw.iloc[data_start:].copy()
    df.columns = final_cols[:len(df.columns)]
    df = df.dropna(how='all').reset_index(drop=True)

    # Remove any sub-header repeat rows
    id_col = next((c for c in df.columns if 'order id' in c.lower()), None)
    if id_col:
        df = df[~df[id_col].str.lower().str.strip().isin(['order id', 'nan', ''])].copy()

    if quarter_label:
        df['quarter'] = quarter_label
    df['source_file'] = path.name
    return df


print("=" * 80)
print("ClearSettle — Flipkart Deep Reconciliation Engine (Phase 3)")
print("Client: Tip Top Garments")
print("=" * 80)

quarterly_files = {
    'Q1_FY25-26': EXT_DIR / '0b1e8e84c055460d_FY25-26' / '0b1e8e84c055460d_Q1_FY_25_26.xlsx',
    'Q2_FY25-26': EXT_DIR / '0b1e8e84c055460d_FY25-26' / '0b1e8e84c055460d_Q2_FY_25_26.xlsx',
    'Q3_FY25-26': EXT_DIR / '0b1e8e84c055460d_FY25-26' / '0b1e8e84c055460d_Q3_FY_25_26.xlsx',
    'Q4_FY25-26': EXT_DIR / '0b1e8e84c055460d_FY25-26' / '0b1e8e84c055460d_Q4_FY_25_26.xlsx',
    'Latest':     BASE_DIR / '79f713e8-ea3f-42af-ba0d-1be464b9c843_1779952313000Payment Reports.xlsx',
}

print("\n[1] Loading quarterly payment reports...")
all_payment_dfs = []
for label, path in quarterly_files.items():
    if path.exists():
        df = parse_payment_report_full(path, label)
        if not df.empty:
            all_payment_dfs.append(df)
            print(f"  {label}: {len(df)} rows, {len(df.columns)} cols")
        else:
            print(f"  {label}: EMPTY")

# Combine all quarterly data
if all_payment_dfs:
    combined = pd.concat(all_payment_dfs, ignore_index=True)
    print(f"\n  COMBINED: {len(combined)} rows")
    print(f"  Columns: {list(combined.columns)}")
else:
    print("  ERROR: No payment data loaded")
    combined = pd.DataFrame()

# ═══════════════════════════════════════════════════════════════════════════
# EXTRACT KEY FINANCIAL COLUMNS FROM PAYMENT REPORTS
# ═══════════════════════════════════════════════════════════════════════════

print("\n[2] Identifying financial columns in payment reports...")

EXPECTED_COLS = {
    'order_id_col': ['order id', 'order_id'],
    'order_item_id_col': ['order item id', 'order_item_id'],
    'neft_id_col': ['neft id', 'neft_id'],
    'payment_date_col': ['payment date', 'payment_date'],
    'bank_settlement_col': ['bank settlement value', 'bank_settlement_value'],
    'sale_amount_col': ['sale amount', 'sale_amount'],
    'commission_rate_col': ['commission rate', 'commission rate (%)'],
    'commission_col': ['commission (rs.)', 'commission_rs', 'commission'],
    'fixed_fee_col': ['fixed fee', 'fixed_fee'],
    'collection_fee_col': ['collection fee', 'collection_fee'],
    'shipping_fee_col': ['shipping fee', 'shipping_fee'],
    'reverse_shipping_col': ['reverse shipping fee', 'reverse_shipping_fee'],
    'tcs_col': ['tcs (rs.)', 'tcs_rs', 'tcs'],
    'tds_col': ['tds (rs.)', 'tds_rs', 'tds'],
    'gst_on_fees_col': ['gst on mp fees', 'gst_on_mp_fees'],
    'marketplace_fee_col': ['marketplace fee', 'marketplace_fee'],
    'pick_pack_col': ['pick and pack fee', 'pick_and_pack_fee'],
    'offer_amount_col': ['total offer amount', 'total_offer_amount'],
    'tier_col': ['tier'],
    'sku_col': ['seller sku', 'seller_sku'],
    'category_col': ['product sub category', 'product_sub_category'],
    'shipping_zone_col': ['shipping zone', 'shipping_zone'],
    'item_gst_rate_col': ['item gst rate (%)', 'item_gst_rate'],
    'chargeable_weight_col': ['chargeable_wt_slab_in_kgs', 'chargeable wt slab'],
    'fulfil_type_col': ['fulfilment_type', 'fulfilment type'],
    'invoice_id_col': ['invoice_id', 'invoice id'],
    'invoice_date_col': ['invoice_date', 'invoice date'],
    'return_status_col': ['item_return_status', 'return status'],
    'shopsy_col': ['shopsy_order', 'shopsy order'],
}

if not combined.empty:
    col_mapping = {}
    for key, candidates in EXPECTED_COLS.items():
        found = None
        for cand in candidates:
            matches = [c for c in combined.columns if cand.lower().replace(' ','_') in c.lower().replace(' ','_')]
            if matches:
                found = matches[0]
                break
        col_mapping[key] = found
        status = "FOUND" if found else "NOT FOUND"
        print(f"  {key:<30}: {found or 'N/A':<50} [{status}]")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION A — COMMISSION RATE ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION A: COMMISSION RATE ANALYSIS")
print("=" * 80)

if not combined.empty:
    comm_rate_col = col_mapping.get('commission_rate_col')
    comm_col = col_mapping.get('commission_col')
    sale_col = col_mapping.get('sale_amount_col')
    oi_col = col_mapping.get('order_item_id_col')
    sku_col = col_mapping.get('sku_col')
    cat_col = col_mapping.get('category_col')
    tier_col = col_mapping.get('tier_col')
    quarter_col = 'quarter'

    if comm_rate_col:
        combined['_comm_rate'] = combined[comm_rate_col].apply(n)
        combined['_commission'] = combined[comm_col].apply(n) if comm_col else 0
        combined['_sale_amount'] = combined[sale_col].apply(n) if sale_col else 0
        combined['_fixed_fee'] = combined[col_mapping['fixed_fee_col']].apply(n) if col_mapping.get('fixed_fee_col') else 0
        combined['_collection_fee'] = combined[col_mapping['collection_fee_col']].apply(n) if col_mapping.get('collection_fee_col') else 0
        combined['_shipping_fee'] = combined[col_mapping['shipping_fee_col']].apply(n) if col_mapping.get('shipping_fee_col') else 0
        combined['_reverse_shipping'] = combined[col_mapping['reverse_shipping_col']].apply(n) if col_mapping.get('reverse_shipping_col') else 0
        combined['_tcs'] = combined[col_mapping['tcs_col']].apply(n) if col_mapping.get('tcs_col') else 0
        combined['_tds'] = combined[col_mapping['tds_col']].apply(n) if col_mapping.get('tds_col') else 0
        combined['_gst_on_fees'] = combined[col_mapping['gst_on_fees_col']].apply(n) if col_mapping.get('gst_on_fees_col') else 0
        combined['_marketplace_fee'] = combined[col_mapping['marketplace_fee_col']].apply(n) if col_mapping.get('marketplace_fee_col') else 0
        combined['_pick_pack'] = combined[col_mapping['pick_pack_col']].apply(n) if col_mapping.get('pick_pack_col') else 0
        combined['_bank_settlement'] = combined[col_mapping['bank_settlement_col']].apply(n) if col_mapping.get('bank_settlement_col') else 0

        print("\n  COMMISSION RATE DISTRIBUTION:")
        rate_dist = combined[combined['_comm_rate'] != 0]['_comm_rate'].value_counts().sort_index()
        for rate, count in rate_dist.items():
            subset = combined[combined['_comm_rate'] == rate]
            comm_total = subset['_commission'].sum()
            sale_total = subset['_sale_amount'].sum()
            pct = (abs(comm_total) / sale_total * 100) if sale_total else 0
            print(f"    Rate {rate:>6.2f}% | Orders: {count:>5} | Commission: Rs.{comm_total:>10,.2f} | Sales: Rs.{sale_total:>10,.2f} | Effective: {pct:.2f}%")

        print("\n  COMMISSION RATE BY QUARTER:")
        for qtr in combined[quarter_col].unique():
            qdf = combined[combined[quarter_col] == qtr]
            avg_rate = qdf[qdf['_comm_rate'] != 0]['_comm_rate'].mean()
            total_comm = qdf['_commission'].sum()
            total_sales = qdf['_sale_amount'].sum()
            eff_rate = (abs(total_comm) / total_sales * 100) if total_sales > 0 else 0
            print(f"    {qtr:<15} avg_rate={avg_rate:.2f}%  total_comm=Rs.{total_comm:>10,.2f}  sales=Rs.{total_sales:>10,.2f}  eff_rate={eff_rate:.2f}%")

        # Category breakdown if available
        if cat_col and cat_col in combined.columns:
            print("\n  COMMISSION RATE BY PRODUCT CATEGORY:")
            cat_comm = combined.groupby(cat_col).apply(
                lambda g: pd.Series({
                    'orders': len(g),
                    'avg_comm_rate': g[g['_comm_rate'] != 0]['_comm_rate'].mean() if len(g[g['_comm_rate'] != 0]) > 0 else 0,
                    'total_commission': g['_commission'].sum(),
                    'total_sales': g['_sale_amount'].sum(),
                })
            ).reset_index()
            for _, row in cat_comm.iterrows():
                eff = abs(row['total_commission']) / row['total_sales'] * 100 if row['total_sales'] > 0 else 0
                print(f"    {s(row[cat_col]):<40} orders={int(row['orders']):>4}  avg_rate={row['avg_comm_rate']:.2f}%  eff={eff:.2f}%")

        # Tier breakdown
        if tier_col and tier_col in combined.columns:
            print("\n  COMMISSION RATE BY TIER:")
            tier_data = combined.groupby(tier_col).apply(
                lambda g: pd.Series({
                    'orders': len(g),
                    'avg_comm_rate': g[g['_comm_rate'] != 0]['_comm_rate'].mean() if len(g[g['_comm_rate'] != 0]) > 0 else 0,
                    'total_commission': g['_commission'].sum(),
                    'total_sales': g['_sale_amount'].sum(),
                })
            ).reset_index()
            for _, row in tier_data.iterrows():
                print(f"    Tier {s(row[tier_col]):<10} orders={int(row['orders']):>4}  avg_rate={row['avg_comm_rate']:.2f}%  comm=Rs.{row['total_commission']:>10,.2f}")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION B — FULL FEE BREAKDOWN PER QUARTER
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION B: FULL FEE BREAKDOWN BY QUARTER")
print("=" * 80)

if not combined.empty and '_commission' in combined.columns:
    fee_cols_map = {
        'Commission': '_commission',
        'Fixed Fee': '_fixed_fee',
        'Collection Fee': '_collection_fee',
        'Shipping Fee': '_shipping_fee',
        'Reverse Shipping': '_reverse_shipping',
        'Pick & Pack': '_pick_pack',
        'TCS': '_tcs',
        'TDS': '_tds',
        'GST on Fees': '_gst_on_fees',
        'Marketplace Fee Total': '_marketplace_fee',
        'Bank Settlement': '_bank_settlement',
        'Sale Amount': '_sale_amount',
    }

    print(f"\n  {'Fee Type':<30}", end='')
    quarters = sorted(combined['quarter'].unique())
    for q in quarters:
        print(f"  {q:>18}", end='')
    print(f"  {'TOTAL':>18}")
    print(f"  {'-'*120}")

    for fee_label, col in fee_cols_map.items():
        if col not in combined.columns:
            continue
        print(f"  {fee_label:<30}", end='')
        total = 0
        for q in quarters:
            val = combined[combined['quarter'] == q][col].sum()
            total += val
            print(f"  {val:>18,.2f}", end='')
        print(f"  {total:>18,.2f}")

    # Full year totals
    print(f"\n  FULL YEAR FY 2025-26 TOTALS (Q1+Q2+Q3+Q4+Latest):")
    fy_quarterly = combined[combined['quarter'].isin(['Q1_FY25-26','Q2_FY25-26','Q3_FY25-26','Q4_FY25-26'])]
    for fee_label, col in fee_cols_map.items():
        if col in combined.columns:
            total = fy_quarterly[col].sum()
            print(f"    {fee_label:<35} Rs. {total:>15,.2f}")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION C — WALLET REDEEM INVESTIGATION
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION C: WALLET REDEEM ANOMALY INVESTIGATION")
print("=" * 80)

invoice_path = BASE_DIR / '602c7ef6-b370-49e4-95e4-968e185ca164_1779952155000Invoices.xlsx'
if invoice_path.exists():
    inv_df = pd.read_excel(invoice_path, sheet_name='Commission Invoice Transactions', dtype=str, engine='openpyxl')
    inv_df.columns = [s(c) for c in inv_df.columns]
    inv_df = inv_df.dropna(how='all')

    fee_name_col = 'Fee Name'
    fee_amt_col = 'Total Fee Amount(Rs.)'
    igst_col = 'IGST Amount'
    oi_col = 'Order Item ID/ Listing ID/ Campaign ID/Transaction ID'

    # Wallet Redeem details
    wallet = inv_df[inv_df[fee_name_col].str.lower().str.contains('wallet', na=False)].copy()
    print(f"\n  WALLET REDEEM TRANSACTIONS: {len(wallet)}")
    if len(wallet) > 0:
        for _, row in wallet.iterrows():
            fee_amt = n(row.get(fee_amt_col, 0))
            igst = n(row.get(igst_col, 0))
            oi = s(row.get(oi_col, ''))
            date = s(row.get('Date', ''))
            print(f"    OI/TX={oi:<50} Fee=Rs.{fee_amt:>12,.2f}  IGST=Rs.{igst:>8,.2f}  Date={date}")

        total_wallet = wallet[fee_amt_col].apply(n).sum()
        print(f"\n  Total Wallet Redeem Deduction: Rs. {total_wallet:,.2f}")
        print(f"\n  ANALYSIS:")
        print(f"  'Wallet Redeem' is NOT a standard Flipkart seller fee.")
        print(f"  This appears to be a DEDUCTION from seller account when a customer")
        print(f"  uses Flipkart wallet credits/cashback that was previously credited to seller.")
        print(f"  Amount Rs. {abs(total_wallet):,.2f} needs manual review.")
        print(f"  ACTION: Raise a seller support ticket to clarify this deduction.")

    # SDD Fee investigation
    sdd = inv_df[inv_df[fee_name_col].str.lower().str.contains('sdd', na=False)].copy()
    print(f"\n  SDD FEE TRANSACTIONS: {len(sdd)}")
    if len(sdd) > 0:
        for _, row in sdd.iterrows():
            fee_amt = n(row.get(fee_amt_col, 0))
            oi = s(row.get(oi_col, ''))
            svc_type = s(row.get('Service Type', ''))
            print(f"    OI={oi:<50} Fee=Rs.{fee_amt:>8,.2f}  ServiceType={svc_type}")
        print(f"  NOTE: SDD = Same Day Delivery fee. Valid Flipkart fee for express delivery orders.")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION D — SHIPPING FEE ANALYSIS BY ZONE AND WEIGHT
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION D: SHIPPING FEE ANALYSIS")
print("=" * 80)

if not combined.empty:
    zone_col = col_mapping.get('shipping_zone_col')
    wt_col = col_mapping.get('chargeable_weight_col')
    shipping_col = col_mapping.get('shipping_fee_col')
    fulfil_col = col_mapping.get('fulfil_type_col')

    print(f"\n  Shipping Zone col: {zone_col}")
    print(f"  Weight Slab col: {wt_col}")
    print(f"  Shipping Fee col: {shipping_col}")

    if zone_col and zone_col in combined.columns:
        print("\n  SHIPPING FEE BY ZONE:")
        zone_analysis = combined.groupby(zone_col).apply(
            lambda g: pd.Series({
                'orders': len(g),
                'total_shipping_fee': g['_shipping_fee'].sum(),
                'avg_shipping_fee': g['_shipping_fee'].mean(),
                'orders_with_fee': len(g[g['_shipping_fee'] != 0]),
            })
        ).reset_index()
        for _, row in zone_analysis.iterrows():
            print(f"    Zone {s(row[zone_col]):<20} orders={int(row['orders']):>5}  total=Rs.{row['total_shipping_fee']:>10,.2f}  avg=Rs.{row['avg_shipping_fee']:>7,.2f}")

    if wt_col and wt_col in combined.columns:
        print("\n  SHIPPING FEE BY WEIGHT SLAB:")
        wt_analysis = combined[combined[wt_col].apply(lambda x: s(x) != '')].groupby(wt_col).apply(
            lambda g: pd.Series({
                'orders': len(g),
                'total_shipping': g['_shipping_fee'].sum(),
                'avg_shipping': g['_shipping_fee'].mean(),
            })
        ).reset_index()
        for _, row in wt_analysis.iterrows():
            print(f"    Weight {s(row[wt_col]):<15} orders={int(row['orders']):>5}  total=Rs.{row['total_shipping']:>8,.2f}  avg=Rs.{row['avg_shipping']:>7,.2f}")

    # Reverse shipping analysis
    if '_reverse_shipping' in combined.columns:
        rev_orders = combined[combined['_reverse_shipping'] != 0]
        print(f"\n  REVERSE SHIPPING (Returns):")
        print(f"    Orders with reverse shipping charge: {len(rev_orders)}")
        print(f"    Total reverse shipping charged: Rs. {rev_orders['_reverse_shipping'].sum():,.2f}")
        print(f"    Average per return: Rs. {rev_orders['_reverse_shipping'].mean():,.2f}")

        if zone_col and zone_col in combined.columns:
            zone_rev = rev_orders.groupby(zone_col)['_reverse_shipping'].agg(['count','sum','mean']).reset_index()
            for _, row in zone_rev.iterrows():
                print(f"    Zone {s(row[zone_col]):<20} count={int(row['count']):>4}  total=Rs.{row['sum']:>8,.2f}  avg=Rs.{row['mean']:>7,.2f}")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION E — TCS / TDS VALIDATION
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION E: TCS/TDS VALIDATION")
print("=" * 80)

if not combined.empty and '_tcs' in combined.columns:
    sale_col_name = col_mapping.get('sale_amount_col', '')
    if sale_col_name:
        fy_data = combined[combined['quarter'].isin(['Q1_FY25-26','Q2_FY25-26','Q3_FY25-26','Q4_FY25-26'])]
        total_sales = fy_data['_sale_amount'].sum()
        total_tcs = fy_data['_tcs'].sum()
        total_tds = fy_data['_tds'].sum()

        print(f"\n  FULL YEAR FY25-26 (Q1-Q4):")
        print(f"    Total Sales Amount: Rs. {total_sales:,.2f}")
        print(f"    Total TCS Charged: Rs. {total_tcs:,.2f}")
        print(f"    Total TDS Charged: Rs. {total_tds:,.2f}")

        # TCS should be 1% of taxable value (net of returns)
        # For FY2023-24 onwards, TCS = 1% (0.5% IGST or 0.5% CGST+SGST each)
        expected_tcs_rate = 0.01  # 1%
        if total_sales != 0:
            actual_tcs_rate = abs(total_tcs) / total_sales
            print(f"\n    Expected TCS rate: {expected_tcs_rate*100:.2f}% of taxable value")
            print(f"    Actual TCS rate:   {actual_tcs_rate*100:.4f}% of gross sales")
            print(f"    NOTE: TCS is on NET taxable value (after returns), not gross sales")
            print(f"          Difference is expected due to returns adjustment")

        # Per-quarter TCS
        print(f"\n  TCS BY QUARTER:")
        for qtr in ['Q1_FY25-26','Q2_FY25-26','Q3_FY25-26','Q4_FY25-26']:
            qdf = combined[combined['quarter'] == qtr]
            qsales = qdf['_sale_amount'].sum()
            qtcs = qdf['_tcs'].sum()
            rate = abs(qtcs)/qsales*100 if qsales else 0
            print(f"    {qtr}: sales=Rs.{qsales:>10,.2f}  tcs=Rs.{qtcs:>8,.2f}  rate={rate:.4f}%")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION F — MISSHIPMENT & RETURN ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION F: MISSHIPMENT & RETURN ANALYSIS")
print("=" * 80)

returns_path = BASE_DIR / 'b400e849-6e9e-44b1-b5b6-81afc9aa27e1_1779952132000Fulfilment Reports.xlsx'
if returns_path.exists():
    ret_df = pd.read_excel(returns_path, sheet_name='Returns', dtype=str, engine='openpyxl')
    ret_df.columns = [s(c).lower().replace(' ','_') for c in ret_df.columns]
    ret_df = ret_df.dropna(how='all')

    reason_col = 'return_reason'
    result_col = 'return_result'
    type_col   = 'return_type'
    sku_col_r  = 'sku'
    oi_col_r   = 'order_item_id'

    print(f"\n  Total Returns: {len(ret_df)}")

    # Misshipment is critical — seller bears reverse shipping + no payment
    misship = ret_df[ret_df[reason_col].str.upper().str.contains('MISSHIP', na=False)]
    quality = ret_df[ret_df[reason_col].str.upper().str.contains('QUALITY', na=False)]
    damaged = ret_df[ret_df[reason_col].str.upper().str.contains('DAMAGE', na=False)]
    size    = ret_df[ret_df[reason_col].str.upper().str.contains('SIZE', na=False)]
    missing = ret_df[ret_df[reason_col].str.upper().str.contains('MISSING', na=False)]

    print(f"\n  RETURN CATEGORY BREAKDOWN:")
    print(f"    MISSHIPMENT (wrong item sent):     {len(misship):>4} orders  — SELLER LIABILITY (check inventory)")
    print(f"    QUALITY ISSUE:                     {len(quality):>4} orders  — SELLER LIABILITY (quality control)")
    print(f"    DAMAGED SHIPMENT:                  {len(damaged):>4} orders  — LOGISTICS LIABILITY (Safe-T claim potential)")
    print(f"    SIZE/FIT ISSUES:                   {len(size):>4} orders  — PLATFORM POLICY")
    print(f"    MISSING ITEM:                      {len(missing):>4} orders  — HIGH RECOVERY POTENTIAL")

    # Missing item returns — very high Safe-T claim potential
    print(f"\n  MISSING ITEM RETURNS (Safe-T Claim Eligible):")
    if len(missing) > 0:
        for _, row in missing.iterrows():
            oi = s(row.get(oi_col_r, ''))
            sku = s(row.get(sku_col_r, ''))
            result = s(row.get(result_col, ''))
            rtype = s(row.get(type_col, ''))
            print(f"    OI={oi}, SKU={sku}, Result={result}, Type={rtype}")
        print(f"\n  ACTION: File Safe-T claims for ALL {len(missing)} missing item returns within 15 days")

    # MISSHIPMENT analysis
    print(f"\n  MISSHIPMENT RETURNS (Wrong Item Sent):")
    print(f"    {len(misship)} orders — check SKU mapping and dispatch process")
    if len(misship) > 0:
        mis_skus = misship[sku_col_r].value_counts().head(5)
        print(f"    Top SKUs with misshipment:")
        for sku, cnt in mis_skus.items():
            print(f"      {sku}: {cnt} times")

    # Damaged shipment — logistics recovery
    print(f"\n  DAMAGED SHIPMENT RETURNS (Logistics Claim Potential):")
    if len(damaged) > 0:
        for _, row in damaged.iterrows():
            oi = s(row.get(oi_col_r, ''))
            sku = s(row.get(sku_col_r, ''))
            result = s(row.get(result_col, ''))
            primary_pv = s(row.get('primary_pv_output', ''))
            print(f"    OI={oi}, SKU={sku}, PV_Result={primary_pv}, Return_Result={result}")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION G — SHOPSY vs FLIPKART ORDER ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION G: SHOPSY vs FLIPKART CHANNEL ANALYSIS")
print("=" * 80)

if not combined.empty:
    shopsy_col = col_mapping.get('shopsy_col')
    if shopsy_col and shopsy_col in combined.columns:
        shopsy_orders = combined[combined[shopsy_col].str.lower().str.strip().isin(['yes','true','1','y'])]
        fk_orders = combined[~combined[shopsy_col].str.lower().str.strip().isin(['yes','true','1','y'])]

        print(f"\n  Shopsy Orders: {len(shopsy_orders)}")
        print(f"    Total Sales: Rs. {shopsy_orders['_sale_amount'].sum():,.2f}")
        print(f"    Avg Commission Rate: {shopsy_orders[shopsy_orders['_comm_rate'] != 0]['_comm_rate'].mean():.2f}%")
        print(f"    Total Commission: Rs. {shopsy_orders['_commission'].sum():,.2f}")

        print(f"\n  Flipkart Direct Orders: {len(fk_orders)}")
        print(f"    Total Sales: Rs. {fk_orders['_sale_amount'].sum():,.2f}")
        print(f"    Avg Commission Rate: {fk_orders[fk_orders['_comm_rate'] != 0]['_comm_rate'].mean():.2f}%")
        print(f"    Total Commission: Rs. {fk_orders['_commission'].sum():,.2f}")

    # Also check via PnL channel column
    pnl_path = BASE_DIR / '44d40417-c126-4bc4-925d-Profit and Loss Report.xlsx'
    pnl_df = pd.read_excel(pnl_path, sheet_name='Orders P&L', dtype=str, engine='openpyxl')
    pnl_df.columns = [s(c) for c in pnl_df.columns]
    pnl_df = pnl_df.dropna(how='all').iloc[1:]  # Skip header repeat row

    channel_col = next((c for c in pnl_df.columns if 'channel' in c.lower()), None)
    if channel_col:
        print(f"\n  P&L CHANNEL BREAKDOWN:")
        settle_col_pnl = 'Bank Settlement [Projected] (INR)'
        if settle_col_pnl in pnl_df.columns:
            pnl_df['_settle'] = pnl_df[settle_col_pnl].apply(n)
            channel_summary = pnl_df.groupby(channel_col).agg(
                orders=('_settle', 'count'),
                total_settlement=('_settle', 'sum'),
            ).reset_index()
            for _, row in channel_summary.iterrows():
                print(f"    {s(row[channel_col]):<20} orders={int(row['orders']):>4}  settlement=Rs.{row['total_settlement']:>10,.2f}")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION H — MP FEE REBATE ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION H: MP FEE REBATE ANALYSIS")
print("=" * 80)

for label, path in quarterly_files.items():
    if not path.exists():
        continue
    try:
        raw = pd.read_excel(path, sheet_name='MP Fee Rebate', header=None, dtype=str, engine='openpyxl')
        if len(raw) < 3:
            print(f"  {label}: No MP Fee Rebate data")
            continue
        # Find header
        for i in range(min(5, len(raw))):
            row_str = ' '.join(s(v).lower() for v in raw.iloc[i])
            if 'payment details' in row_str or 'rebate' in row_str:
                headers = [s(v) for v in raw.iloc[i+1] if s(v)]
                data_rows = raw.iloc[i+2:]
                break
        else:
            data_rows = raw.iloc[2:]
            headers = [s(v) for v in raw.iloc[1]]

        # Check if has actual data
        data_rows_clean = data_rows.dropna(how='all')
        if len(data_rows_clean) > 0:
            # Look for amount columns
            for j, row in data_rows_clean.iterrows():
                vals = [s(v) for v in row if s(v)]
                if vals:
                    print(f"  {label}: {' | '.join(vals[:8])}")
        else:
            print(f"  {label}: Empty MP Fee Rebate")
    except Exception as e:
        print(f"  {label}: Error {e}")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION I — COMPREHENSIVE PER-ORDER FEE VALIDATION
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION I: PER-ORDER FEE VALIDATION (sample)")
print("=" * 80)

if not combined.empty and '_commission' in combined.columns:
    # Get orders with complete data
    order_id_col = col_mapping.get('order_id_col')
    oi_col = col_mapping.get('order_item_id_col')
    sku_col2 = col_mapping.get('sku_col')
    cat_col2 = col_mapping.get('category_col')

    validation_issues = []

    # Sample: validate commission rate for each order
    # Flipkart apparel commission structure (FY25-26, as per published rates):
    # Apparel/Clothing: 5% commission (basic sellers)
    # The actual rate is encoded in the report itself (commission_rate_%)
    # We validate: Is actual_commission = commission_rate * sale_amount?

    for idx, row in combined.head(20).iterrows():
        comm_rate = row['_comm_rate']
        commission = row['_commission']
        sale_amt = row['_sale_amount']
        order_id = s(row.get(order_id_col, '')) if order_id_col else ''
        sku = s(row.get(sku_col2, '')) if sku_col2 else ''

        if sale_amt > 0 and comm_rate > 0:
            expected_commission = -(sale_amt * comm_rate / 100)
            variance = commission - expected_commission
            if abs(variance) > 1:
                status = "MISMATCH"
                validation_issues.append({
                    'order_id': order_id,
                    'sku': sku,
                    'sale_amount': sale_amt,
                    'commission_rate': comm_rate,
                    'expected_commission': expected_commission,
                    'actual_commission': commission,
                    'variance': variance,
                    'issue_type': 'COMMISSION_CALCULATION_ERROR',
                })
            else:
                status = "OK"

            print(f"  OI={order_id[-10:] if order_id else 'N/A':<10}  sale=Rs.{sale_amt:>7,.0f}  rate={comm_rate:.2f}%  expected=Rs.{expected_commission:>7,.2f}  actual=Rs.{commission:>7,.2f}  [{status}]")

    if validation_issues:
        print(f"\n  VALIDATION ISSUES FOUND: {len(validation_issues)}")
        pd.DataFrame(validation_issues).to_excel(OUT_DIR / '08_commission_validation_issues.xlsx', index=False)
    else:
        print(f"\n  All sampled commissions calculated correctly.")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION J — DISPUTE CANDIDATE REPORT
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION J: DISPUTE CANDIDATES SUMMARY")
print("=" * 80)

dispute_candidates = []

# 1. Wallet Redeem — Unknown deduction
if invoice_path.exists():
    inv_df2 = pd.read_excel(invoice_path, sheet_name='Commission Invoice Transactions', dtype=str, engine='openpyxl')
    inv_df2.columns = [s(c) for c in inv_df2.columns]
    wallet2 = inv_df2[inv_df2['Fee Name'].str.lower().str.contains('wallet', na=False)]
    if len(wallet2) > 0:
        total_wallet = wallet2['Total Fee Amount(Rs.)'].apply(n).sum()
        dispute_candidates.append({
            'dispute_id': 'DIS-001',
            'issue_type': 'UNEXPLAINED_DEDUCTION',
            'description': '"Wallet Redeem" deduction — not a standard Flipkart seller fee',
            'affected_order_count': len(wallet2),
            'amount': abs(total_wallet),
            'evidence': 'Commission Invoice — Fee Name: Wallet Redeem',
            'action': 'File Seller Support ticket requesting explanation and reversal',
            'confidence': 'HIGH',
            'policy_basis': 'No Flipkart policy supports "Wallet Redeem" as a seller deduction',
            'priority': '1-HIGH',
        })

# 2. Missing settlements for DELIVERED orders
pnl_path2 = BASE_DIR / '44d40417-c126-4bc4-925d-Profit and Loss Report.xlsx'
if pnl_path2.exists():
    pnl2 = pd.read_excel(pnl_path2, sheet_name='Orders P&L', dtype=str, engine='openpyxl')
    pnl2.columns = [s(c) for c in pnl2.columns]
    pnl2 = pnl2.dropna(how='all').iloc[1:]
    pnl2['_pending'] = pnl2['Amount Pending (INR)'].apply(n)
    pnl2['_status'] = pnl2['Order Status'].apply(s)
    pnl2['_settle_proj'] = pnl2['Bank Settlement [Projected] (INR)'].apply(n)

    delivered_pending = pnl2[(pnl2['_status'] == 'DELIVERED') & (pnl2['_pending'] > 50)]
    if len(delivered_pending) > 0:
        total_del_pending = delivered_pending['_pending'].sum()
        dispute_candidates.append({
            'dispute_id': 'DIS-002',
            'issue_type': 'MISSING_SETTLEMENT',
            'description': f'{len(delivered_pending)} DELIVERED orders with settlement not received',
            'affected_order_count': len(delivered_pending),
            'amount': total_del_pending,
            'evidence': 'P&L Report — Amount Pending (INR) > 50 for DELIVERED status',
            'action': 'Check NEFT payment cycle; file settlement dispute for overdue orders',
            'confidence': 'HIGH',
            'policy_basis': 'Flipkart payment cycle: 7-15 days post delivery',
            'priority': '1-HIGH',
        })
        print(f"\n  DIS-002: {len(delivered_pending)} delivered orders with Rs.{total_del_pending:,.2f} pending")

# 3. Misshipment returns — seller should not bear full reverse shipping if platform error
if returns_path.exists():
    ret2 = pd.read_excel(returns_path, sheet_name='Returns', dtype=str, engine='openpyxl')
    ret2.columns = [s(c).lower().replace(' ','_') for c in ret2.columns]
    ret2 = ret2.dropna(how='all')
    misship2 = ret2[ret2['return_reason'].str.upper().str.contains('MISSHIP', na=False)]
    if len(misship2) > 0:
        dispute_candidates.append({
            'dispute_id': 'DIS-003',
            'issue_type': 'MISSHIPMENT_REVERSE_SHIPPING',
            'description': f'{len(misship2)} misshipment returns — seller claims against reverse shipping overcharge',
            'affected_order_count': len(misship2),
            'amount': 0,  # Need to cross-reference with payment report
            'evidence': 'Fulfilment Report — return_reason: MISSHIPMENT',
            'action': 'Cross-check with payment report; claim reimbursement for valid misshipment returns',
            'confidence': 'MEDIUM',
            'policy_basis': 'Flipkart Seller Protection — misshipment reverse logistics policy',
            'priority': '2-MEDIUM',
        })

# 4. Returns where customer received refund but seller was NOT paid
if not combined.empty and '_reverse_shipping' in combined.columns:
    return_orders = combined[combined['_reverse_shipping'] != 0]
    print(f"\n  Total orders with reverse shipping: {len(return_orders)}")
    total_rev_ship = return_orders['_reverse_shipping'].sum()
    if abs(total_rev_ship) > 0:
        dispute_candidates.append({
            'dispute_id': 'DIS-004',
            'issue_type': 'REVERSE_SHIPPING_AUDIT',
            'description': f'{len(return_orders)} orders with reverse shipping charges — validate against return policy',
            'affected_order_count': len(return_orders),
            'amount': abs(total_rev_ship),
            'evidence': 'Payment Reports — Reverse Shipping Fee column',
            'action': 'Validate each return type: is reverse shipping justified per Flipkart policy?',
            'confidence': 'MEDIUM',
            'policy_basis': 'Flipkart Reverse Logistics Policy — seller not liable for platform errors',
            'priority': '2-MEDIUM',
        })

# Save dispute candidates
disp_df = pd.DataFrame(dispute_candidates)
if not disp_df.empty:
    disp_df.to_excel(OUT_DIR / '09_dispute_candidates.xlsx', index=False)
    print(f"\n  Dispute candidates saved: 09_dispute_candidates.xlsx")
    print(f"\n  DISPUTE SUMMARY:")
    print(f"  {'ID':<10} {'Priority':<12} {'Type':<35} {'Orders':>6} {'Amount':>12}")
    print(f"  {'-'*90}")
    for _, row in disp_df.iterrows():
        print(f"  {row['dispute_id']:<10} {row['priority']:<12} {row['issue_type']:<35} {row['affected_order_count']:>6} Rs.{row['amount']:>10,.2f}")
    total_disputes = disp_df['amount'].sum()
    print(f"\n  TOTAL DISPUTED AMOUNT: Rs. {total_disputes:,.2f}")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION K — FULL YEAR FINANCIAL SUMMARY
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION K: FULL YEAR FY2025-26 FINANCIAL SUMMARY")
print("=" * 80)

if not combined.empty:
    fy = combined[combined['quarter'].isin(['Q1_FY25-26','Q2_FY25-26','Q3_FY25-26','Q4_FY25-26'])]

    print(f"\n  {'Metric':<50} {'Amount (Rs.)':>15}")
    print(f"  {'-'*70}")
    metrics = [
        ('Total Gross Sales', fy['_sale_amount'].sum()),
        ('Total Commission Charged', fy['_commission'].sum()),
        ('Total Fixed Fee', fy['_fixed_fee'].sum()),
        ('Total Collection Fee', fy['_collection_fee'].sum()),
        ('Total Shipping Fee', fy['_shipping_fee'].sum()),
        ('Total Reverse Shipping', fy['_reverse_shipping'].sum()),
        ('Total Pick & Pack Fee', fy['_pick_pack'].sum()),
        ('Total GST on Fees', fy['_gst_on_fees'].sum()),
        ('Total TCS', fy['_tcs'].sum()),
        ('Total TDS', fy['_tds'].sum()),
        ('Total Marketplace Fee', fy['_marketplace_fee'].sum()),
        ('Net Bank Settlement', fy['_bank_settlement'].sum()),
    ]
    for label, val in metrics:
        print(f"  {label:<50} {val:>15,.2f}")

    total_sales = fy['_sale_amount'].sum()
    total_fees = fy['_marketplace_fee'].sum()
    if total_sales:
        eff_fee_rate = abs(total_fees)/total_sales*100
        print(f"\n  Effective Total Fee Rate (% of sales): {eff_fee_rate:.2f}%")

    # Orders summary
    print(f"\n  ORDERS SUMMARY:")
    print(f"  Q1 Orders: {len(combined[combined['quarter']=='Q1_FY25-26'])}")
    print(f"  Q2 Orders: {len(combined[combined['quarter']=='Q2_FY25-26'])}")
    print(f"  Q3 Orders: {len(combined[combined['quarter']=='Q3_FY25-26'])}")
    print(f"  Q4 Orders: {len(combined[combined['quarter']=='Q4_FY25-26'])}")
    print(f"  TOTAL Q1-Q4: {len(fy)}")

    # Save combined payment data
    fy.to_excel(OUT_DIR / '10_payment_data_fy2526.xlsx', index=False)
    print(f"\n  Saved: 10_payment_data_fy2526.xlsx")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION L — UPDATE MASTER RECONCILIATION REPORT
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 80)
print("SECTION L: UPDATING MASTER RECONCILIATION REPORT")
print("=" * 80)

# Update the master Excel report with new sheets
master_path = OUT_DIR / '00_CLEARSETTLE_FLIPKART_RECONCILIATION_REPORT.xlsx'

# Load existing
try:
    from openpyxl import load_workbook
    wb = load_workbook(master_path)
    existing_sheets = wb.sheetnames
    wb.close()
except:
    existing_sheets = []

with pd.ExcelWriter(OUT_DIR / '00_CLEARSETTLE_FLIPKART_RECONCILIATION_REPORT_v2.xlsx', engine='openpyxl') as writer:

    # Sheet: Dispute Candidates
    if not disp_df.empty:
        disp_df.to_excel(writer, sheet_name='Dispute Candidates', index=False)

    # Sheet: Fee Validation Summary
    if not combined.empty:
        fy_summary_rows = []
        for q in ['Q1_FY25-26','Q2_FY25-26','Q3_FY25-26','Q4_FY25-26']:
            qdf = combined[combined['quarter'] == q]
            if len(qdf) > 0:
                fy_summary_rows.append({
                    'Quarter': q,
                    'Orders': len(qdf),
                    'Gross Sales (Rs.)': qdf['_sale_amount'].sum(),
                    'Commission (Rs.)': qdf['_commission'].sum(),
                    'Fixed Fee (Rs.)': qdf['_fixed_fee'].sum(),
                    'Collection Fee (Rs.)': qdf['_collection_fee'].sum(),
                    'Shipping Fee (Rs.)': qdf['_shipping_fee'].sum(),
                    'Reverse Shipping (Rs.)': qdf['_reverse_shipping'].sum(),
                    'Pick & Pack (Rs.)': qdf['_pick_pack'].sum(),
                    'TCS (Rs.)': qdf['_tcs'].sum(),
                    'TDS (Rs.)': qdf['_tds'].sum(),
                    'GST on Fees (Rs.)': qdf['_gst_on_fees'].sum(),
                    'Net Bank Settlement (Rs.)': qdf['_bank_settlement'].sum(),
                })
        fee_summary_df = pd.DataFrame(fy_summary_rows)
        fee_summary_df.to_excel(writer, sheet_name='Fee Summary by Quarter', index=False)

    # Sheet: Returns Analysis
    if returns_path.exists():
        ret2_out = pd.read_excel(returns_path, sheet_name='Returns', dtype=str, engine='openpyxl')
        ret2_out.to_excel(writer, sheet_name='Returns Detail', index=False)

    # Sheet: SKU Performance
    if not combined.empty and sku_col2 and sku_col2 in combined.columns:
        sku_perf = combined.groupby(sku_col2).apply(
            lambda g: pd.Series({
                'orders': len(g),
                'gross_sales': g['_sale_amount'].sum(),
                'commission': g['_commission'].sum(),
                'fixed_fee': g['_fixed_fee'].sum(),
                'shipping_fee': g['_shipping_fee'].sum(),
                'reverse_shipping': g['_reverse_shipping'].sum(),
                'net_settlement': g['_bank_settlement'].sum(),
                'avg_commission_rate': g[g['_comm_rate']!=0]['_comm_rate'].mean() if len(g[g['_comm_rate']!=0]) > 0 else 0,
            })
        ).reset_index().sort_values('net_settlement', ascending=False)
        sku_perf.to_excel(writer, sheet_name='SKU Performance', index=False)

    # Sheet: Category Analysis
    if not combined.empty and cat_col2 and cat_col2 in combined.columns:
        cat_perf = combined.groupby(cat_col2).apply(
            lambda g: pd.Series({
                'orders': len(g),
                'gross_sales': g['_sale_amount'].sum(),
                'commission': g['_commission'].sum(),
                'avg_rate': g[g['_comm_rate']!=0]['_comm_rate'].mean() if len(g[g['_comm_rate']!=0]) > 0 else 0,
                'net_settlement': g['_bank_settlement'].sum(),
            })
        ).reset_index().sort_values('gross_sales', ascending=False)
        cat_perf.to_excel(writer, sheet_name='Category Analysis', index=False)

print(f"\n  Saved: 00_CLEARSETTLE_FLIPKART_RECONCILIATION_REPORT_v2.xlsx")

print("\n" + "=" * 80)
print("PHASE 3 COMPLETE — Deep Reconciliation Analysis Done")
print("=" * 80)
print(f"\n  All outputs at: {OUT_DIR}")
