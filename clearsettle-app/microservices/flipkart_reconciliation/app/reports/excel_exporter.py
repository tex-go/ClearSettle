from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
_HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(bold=True, size=13, color="1F3864")

_STATUS_FILL = {
    "MATCHED":            PatternFill("solid", fgColor="C6EFCE"),
    "SHORT_PAID":         PatternFill("solid", fgColor="FFCCCC"),
    "OVER_PAID":          PatternFill("solid", fgColor="FFEB9C"),
    "RETURN_RECOVERY":    PatternFill("solid", fgColor="DAE8FC"),
    "MISSING_SETTLEMENT": PatternFill("solid", fgColor="F2DCDB"),
    "MISSING_ORDER":      PatternFill("solid", fgColor="DDDDDD"),
    "MISSING_FEE_RECORD": PatternFill("solid", fgColor="FFF2CC"),
}


def _fmt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, (date, datetime)):
        return str(val)
    return val


def generate_april_report(rows: list[dict], period_label: str = "April 2026") -> bytes:
    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum["A1"] = f"Flipkart Reconciliation Report — {period_label}"
    ws_sum["A1"].font = _TITLE_FONT
    ws_sum["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws_sum["A2"].font = Font(italic=True, size=9, color="666666")

    status_counts: dict[str, int] = {}
    status_diffs: dict[str, Decimal] = {}
    for r in rows:
        s = r["reconciliation_status"]
        status_counts[s] = status_counts.get(s, 0) + 1
        diff = r.get("difference") or Decimal("0")
        status_diffs[s] = status_diffs.get(s, Decimal("0")) + diff

    net_leakage = abs(float(status_diffs.get("SHORT_PAID", Decimal("0"))))
    net_overpaid = float(status_diffs.get("OVER_PAID", Decimal("0")))

    summary_rows = [
        ("Status", "Count", "Net Difference (Rs.)", "Interpretation"),
        ("MATCHED",            status_counts.get("MATCHED", 0),            "0.00",                          "Fee matches invoice exactly"),
        ("SHORT_PAID",         status_counts.get("SHORT_PAID", 0),         float(status_diffs.get("SHORT_PAID", 0)),      "Flipkart deducted more fees than invoiced"),
        ("OVER_PAID",          status_counts.get("OVER_PAID", 0),          float(status_diffs.get("OVER_PAID", 0)),       "Flipkart deducted fewer fees than invoiced"),
        ("RETURN_RECOVERY",    status_counts.get("RETURN_RECOVERY", 0),    float(status_diffs.get("RETURN_RECOVERY", 0)), "Return/refund recovery — not a leakage"),
        ("MISSING_SETTLEMENT", status_counts.get("MISSING_SETTLEMENT", 0), "—",                             "Order not yet settled"),
        ("MISSING_FEE_RECORD", status_counts.get("MISSING_FEE_RECORD", 0), "—",                             "Settlement has fees but no invoice entry"),
        ("MISSING_ORDER",      status_counts.get("MISSING_ORDER", 0),      "—",                             "No order/settlement record"),
        ("TOTAL",              len(rows),                                   "",                              ""),
        ("NET LEAKAGE (Short-paid only)", "",  abs(net_leakage),    "Amount Flipkart owes seller"),
        ("NET OVER-PAID",                 "",  net_overpaid,         "Amount settled above invoice — verify with Flipkart"),
    ]

    start_row = 4
    for i, row_data in enumerate(summary_rows):
        row = start_row + i
        for col, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=row, column=col, value=val)
            if i == 0:
                cell.fill = _HDR_FILL
                cell.font = _HDR_FONT
            elif row_data[0] in _STATUS_FILL:
                cell.fill = _STATUS_FILL[row_data[0]]
        if row_data[0] in ("NET LEAKAGE (Short-paid only)", "TOTAL", "NET OVER-PAID"):
            for col in range(1, 5):
                ws_sum.cell(row=row, column=col).font = Font(bold=True)

    ws_sum.column_dimensions["A"].width = 36
    ws_sum.column_dimensions["B"].width = 10
    ws_sum.column_dimensions["C"].width = 22
    ws_sum.column_dimensions["D"].width = 45

    # ── Detail sheet ───────────────────────────────────────────────────────────
    ws_det = wb.create_sheet("Detail")

    headers = [
        "Order Item ID", "Order ID", "SKU", "Product Title",
        "Order Date", "Settlement Date",
        "Sale Amount (Rs.)",
        "Total Offer Amount (Rs.)", "My Share (Rs.)",
        "Invoice Fee (Rs.)", "Invoice GST (Rs.)",
        "Commission Fee", "Shipping Fee", "Other Fee",
        "TCS (Rs.)", "TDS (Rs.)", "GST on MP Fees",
        "Expected Settlement (Rs.)", "Actual Settlement (Rs.)", "Difference (Rs.)",
        "Status", "NEFT ID", "NEFT Type",
    ]
    field_keys = [
        "order_item_id", "order_id", "sku", "product_title",
        "order_date", "settlement_date",
        "sale_amount",
        "total_offer_amount", "my_share",
        "invoice_fee_total", "invoice_gst_total",
        "commission_fee", "shipping_fee", "other_fee",
        "tcs", "tds", "gst_on_mp_fees",
        "expected_settlement", "settlement_amount", "difference",
        "reconciliation_status", "neft_id", "neft_type",
    ]

    for col, hdr in enumerate(headers, 1):
        cell = ws_det.cell(row=1, column=col, value=hdr)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(rows, 2):
        status = row.get("reconciliation_status", "")
        fill = _STATUS_FILL.get(status)
        for c_idx, key in enumerate(field_keys, 1):
            val = _fmt(row.get(key))
            cell = ws_det.cell(row=r_idx, column=c_idx, value=val)
            if fill and field_keys[c_idx - 1] == "reconciliation_status":
                cell.fill = fill
            # Highlight difference column red/green
            if key == "difference" and isinstance(val, (int, float)):
                if val < -0.5:
                    cell.font = Font(color="CC0000", bold=True)
                elif val > 0.5:
                    cell.font = Font(color="006100", bold=True)

    col_widths = [22, 22, 18, 40, 12, 14, 14, 18, 14, 16, 14, 14, 13, 12, 12, 12, 14, 20, 20, 14, 20, 30, 12]
    for col, width in enumerate(col_widths, 1):
        ws_det.column_dimensions[get_column_letter(col)].width = width

    ws_det.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
