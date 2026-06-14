"""
Flipkart ETL Data Integrity Audit — 8-Rule Validation Suite

Independent validation of the pipeline:
    Excel File → raw tables (orders_raw / fees_raw / settlements_raw)
                → ETL tables  (orders_etl / fees_etl / settlements_etl)

Design principle: NO imports from app.* — all transformations are
re-implemented here from scratch so we get a genuinely independent check.
"""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pandas as pd
import psycopg2
import psycopg2.extras

# ─────────────────────────────────────────────────────────────────────────────
# Column maps (independent copy — deliberately NOT importing from ingestor)
# ─────────────────────────────────────────────────────────────────────────────

ORDERS_COL_MAP: dict[str, str] = {
    # Space-separated variants (Flipkart web downloads)
    "order item id": "order_item_id",
    "order id": "order_id",
    "product title": "product_title",
    "order date": "order_date",
    "order item status": "order_item_status",
    "dispatch date": "dispatch_date",
    "dispatched date": "dispatch_date",
    "order delivery date": "delivery_date",
    "delivery date": "delivery_date",
    "order approval date": "dispatch_date",
    # Underscore variants (Flipkart bulk-download CSVs and some Excel downloads)
    "order_item_id": "order_item_id",
    "order_id": "order_id",
    "sku": "sku",
    "fsn": "fsn",
    "product_title": "product_title",
    "quantity": "quantity",
    "order_date": "order_date",
    "order_item_status": "order_item_status",
    "dispatched_date": "dispatch_date",
    "order_delivery_date": "delivery_date",    # actual Excel column name (underscore)
    "order_approval_date": "dispatch_date",    # actual Excel column name (underscore)
}

FEES_COL_MAP: dict[str, str] = {
    "order item id/ listing id/ campaign id/transaction id": "order_item_id",
    "order item id": "order_item_id",
    "order_item_id": "order_item_id",
    "fee name": "fee_name",
    "fee_name": "fee_name",
    "fee amount (rs.)": "fee_amount",
    "total fee amount(rs.)": "fee_amount",
    "fee_amount": "fee_amount",
    "fee waiver amount(rs.)": "fee_waiver_amount",
    "fee_waiver_amount": "fee_waiver_amount",
    "cgst amount": "cgst",
    "cgst": "cgst",
    "sgst/utgst amount": "sgst",
    "sgst": "sgst",
    "igst amount": "igst",
    "igst": "igst",
    "total tax amount (rs.)": "tax_amount",
    "tax amount": "tax_amount",
    "tax_amount": "tax_amount",
    "date": "fee_date",
    "transaction date": "fee_date",
    "invoice date": "fee_date",
    "fee_date": "fee_date",
}

SETTLEMENTS_COL_MAP: dict[str, str] = {
    "neft id": "neft_id",
    "neft_id": "neft_id",
    "neft reference id": "neft_id",
    "neft type": "neft_type",
    "payment date": "settlement_date",
    "settlement date": "settlement_date",
    "settlement_date": "settlement_date",
    "order id": "order_id",
    "order_id": "order_id",
    "order item id": "order_item_id",
    "order_item_id": "order_item_id",
    "sale amount (rs.)": "sale_amount",
    "selling price": "sale_amount",
    "selling_price": "sale_amount",
    "total offer amount (rs.)": "total_offer_amount",
    "my share (rs.)": "my_share",
    "customer add-ons amount (rs.)": "customer_addons_amount",
    "offer adjustments (rs.)": "offer_adjustments",
    "protection fund (rs.)": "protection_fund",
    "refund (rs.)": "refund",
    "tcs (rs.)": "tcs",
    "tds (rs.)": "tds",
    "gst on mp fees (rs.)": "gst_on_mp_fees",
    "fixed fee (rs.)": "fixed_fee",
    "collection fee (rs.)": "collection_fee",
    "reverse shipping fee (rs.)": "reverse_shipping_fee",
    "seller sku": "sku",
    "sku": "sku",
}

_SETTLEMENT_AMOUNT_PREFIX = "bank settlement value"
_MARKETPLACE_FEE_PREFIX = "marketplace fee"

ORDERS_SHEET_HINTS    = ["orders", "order", "fulfilment", "fulfillment"]
FEES_SHEET_HINTS      = ["commission invoice transactions", "commission", "fee", "invoice", "charges"]
SETTLEMENTS_SHEET_HINTS = ["settlement", "settled", "payment", "payout", "orders"]

ORDERS_KEYWORDS      = ["order item id", "order id", "sku", "status"]
FEES_KEYWORDS        = ["order item id", "fee", "commission", "cgst", "sgst"]
SETTLEMENTS_KEYWORDS = ["neft", "order item id", "order id", "sale amount", "settlement"]

# ETL numeric columns per table
ORDERS_DATE_COLS   = ["order_date", "dispatch_date", "delivery_date"]
FEES_DATE_COLS     = ["fee_date"]
FEES_NUMERIC_COLS  = ["fee_amount", "fee_waiver_amount", "cgst", "sgst", "igst", "tax_amount"]
SETTLEMENTS_DATE_COLS    = ["settlement_date"]
SETTLEMENTS_NUMERIC_COLS = [
    "settlement_amount", "sale_amount", "total_offer_amount", "my_share",
    "customer_addons_amount", "marketplace_fee", "offer_adjustments",
    "protection_fund", "refund", "tcs", "tds", "gst_on_mp_fees",
    "fixed_fee", "collection_fee", "reverse_shipping_fee",
]


# ─────────────────────────────────────────────────────────────────────────────
# Result data classes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class RuleResult:
    rule_id: str
    rule_name: str
    file_type: str          # "orders" | "fees" | "settlements"
    passed: bool
    summary: str
    details: dict[str, Any] = field(default_factory=dict)
    issues: list[dict] = field(default_factory=list)

    @property
    def status(self) -> str:
        return "PASS" if self.passed else "FAIL"


@dataclass
class AuditReport:
    results: list[RuleResult] = field(default_factory=list)

    def add(self, r: RuleResult) -> None:
        self.results.append(r)

    def passed_all(self) -> bool:
        return all(r.passed for r in self.results)

    def failed(self) -> list[RuleResult]:
        return [r for r in self.results if not r.passed]

    def print_summary(self) -> None:
        print("\n" + "=" * 72)
        print("  FLIPKART ETL DATA INTEGRITY AUDIT -- RESULTS")
        print("=" * 72)
        for r in self.results:
            icon = "+" if r.passed else "X"
            print(f"  {icon} [{r.rule_id}] {r.rule_name:<40} [{r.file_type:<12}] {r.status}")
            if not r.passed:
                print(f"      -> {r.summary}")
        print("-" * 72)
        total = len(self.results)
        passed = sum(1 for r in self.results if r.passed)
        print(f"  TOTAL: {passed}/{total} rules passed")
        print("=" * 72 + "\n")


# ─────────────────────────────────────────────────────────────────────────────
# Independent utility functions (re-implemented — no app.* imports)
# ─────────────────────────────────────────────────────────────────────────────

def _clean_col(col: str) -> str:
    """Mirrors ingestor._clean_col_name exactly."""
    return " ".join(str(col).replace("\n", " ").split()).lower()


def _clean_val(value) -> str | None:
    """Mirrors ingestor._clean_val exactly."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    s = str(value).strip().strip('"')
    return s if s else None


def _parse_decimal(value) -> Decimal:
    """Mirrors validator.parse_decimal exactly."""
    if value is None:
        return Decimal("0")
    cleaned = str(value).replace(",", "").replace("₹", "").strip()
    if not cleaned or cleaned.lower() in ("none", "nan", "-", ""):
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0")


def _parse_decimal_optional(value) -> Decimal | None:
    """Mirrors validator.parse_optional_decimal exactly."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("none", "nan", "-", ""):
        return None
    return _parse_decimal(value)


def _parse_date(value) -> date | None:
    """
    Mirrors validator.parse_date but also handles datetime strings
    (e.g. "2026-04-01 16:00:23") by stripping the time component.
    """
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    if not s or s.lower() in ("none", "nan", "nat"):
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S",
                "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Last resort: if there's a space, try the date part alone
    if " " in s:
        return _parse_date(s.split(" ")[0])
    return None


# ─────────────────────────────────────────────────────────────────────────────
# ExcelReader — reads Excel exactly as the ingestor would
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ExcelReadResult:
    file_path: str
    sheet: str
    header_row: int
    raw_columns: list[str]
    mapped_columns: dict[str, str]   # excel_col → canonical_name
    unmapped_columns: list[str]      # cols not in any col_map
    sub_header_dropped: bool
    empty_rows_dropped: int
    df: pd.DataFrame                 # final DataFrame with canonical column names
    df_raw: pd.DataFrame             # DataFrame before column rename (raw col names)


class ExcelReader:
    """
    Reads an Excel file using the exact same logic as the ingestor,
    independently reimplemented for audit purposes.
    """

    def _pick_sheet(self, xl: pd.ExcelFile, hints: list[str]) -> str:
        for hint in hints:
            for name in xl.sheet_names:
                if hint in name.lower():
                    return name
        return xl.sheet_names[0]

    def _find_header_row(self, xl: pd.ExcelFile, sheet: str, keywords: list[str]) -> int:
        probe = pd.read_excel(xl, sheet_name=sheet, header=None, dtype=str, nrows=15)
        for i, row in probe.iterrows():
            row_str = " ".join(str(v).lower() for v in row.values if pd.notna(v))
            if sum(1 for kw in keywords if kw in row_str) >= 2:
                return i
        return 0

    def _build_full_map(self, raw_cols: list[str], base_map: dict[str, str]) -> dict[str, str]:
        resolved = dict(base_map)
        for col in raw_cols:
            if col not in resolved:
                if col.startswith(_SETTLEMENT_AMOUNT_PREFIX):
                    resolved[col] = "settlement_amount"
                elif col.startswith(_MARKETPLACE_FEE_PREFIX):
                    resolved[col] = "marketplace_fee"
        return resolved

    def read(
        self,
        file_path: Path,
        sheet_hints: list[str],
        keywords: list[str],
        col_map: dict[str, str],
    ) -> ExcelReadResult:
        xl = pd.ExcelFile(str(file_path))
        sheet = self._pick_sheet(xl, sheet_hints)
        header_row = self._find_header_row(xl, sheet, keywords)

        df = pd.read_excel(xl, sheet_name=sheet, header=header_row, dtype=str)
        raw_cols = [_clean_col(str(c)) for c in df.columns]
        df.columns = raw_cols

        # Detect and drop sub-header row (mirrors ingestor._read_excel)
        sub_header_dropped = False
        if len(df) > 0:
            first_val = str(df.iloc[0, 0]).strip().lower()
            if "total" in first_val or first_val in ("nan", ""):
                df = df.iloc[1:].reset_index(drop=True)
                sub_header_dropped = True

        rows_before = len(df)
        df_raw = df.copy()
        df = df.dropna(how="all").reset_index(drop=True)
        df_raw = df_raw.dropna(how="all").reset_index(drop=True)
        empty_dropped = rows_before - len(df)

        full_map = self._build_full_map(raw_cols, col_map)
        mapped = {c: full_map[c] for c in raw_cols if c in full_map}
        unmapped = [c for c in raw_cols if c not in full_map]

        df = df.rename(columns=full_map)

        return ExcelReadResult(
            file_path=str(file_path),
            sheet=sheet,
            header_row=header_row,
            raw_columns=raw_cols,
            mapped_columns=mapped,
            unmapped_columns=unmapped,
            sub_header_dropped=sub_header_dropped,
            empty_rows_dropped=empty_dropped,
            df=df,
            df_raw=df_raw,
        )


# ─────────────────────────────────────────────────────────────────────────────
# DatabaseClient — reads tables via psycopg2 (independent of app ORM)
# ─────────────────────────────────────────────────────────────────────────────

class DbClient:
    def __init__(self, dsn: str):
        self.conn = psycopg2.connect(dsn)
        self.conn.autocommit = True

    def close(self) -> None:
        self.conn.close()

    def _query(self, sql: str, params: tuple = ()) -> list[dict]:
        with self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            return [dict(r) for r in cur.fetchall()]

    def latest_batch(self, table: str) -> dict | None:
        rows = self._query(
            f"SELECT batch_id, file_name, COUNT(*) AS row_count, MAX(uploaded_at) AS uploaded_at "
            f"FROM {table} GROUP BY batch_id, file_name "
            f"ORDER BY MAX(uploaded_at) DESC LIMIT 1"
        )
        return rows[0] if rows else None

    def all_batches(self, table: str) -> list[dict]:
        return self._query(
            f"SELECT batch_id, file_name, COUNT(*) AS row_count, MAX(uploaded_at) AS uploaded_at "
            f"FROM {table} GROUP BY batch_id, file_name "
            f"ORDER BY MAX(uploaded_at) DESC"
        )

    def best_matching_batch(self, table: str, target_row_count: int) -> dict | None:
        """
        Finds the batch whose row_count is closest to target_row_count.
        Prefer exact match; fall back to nearest. This lets the audit correctly
        match an April Excel (529 rows) to the April DB batch even when a May
        batch (763 rows) is the most recently uploaded.
        """
        batches = self.all_batches(table)
        if not batches:
            return None
        exact = [b for b in batches if b["row_count"] == target_row_count]
        if exact:
            return exact[0]   # prefer most-recent exact match
        return min(batches, key=lambda b: abs(b["row_count"] - target_row_count))

    def batches_for_file(self, table: str, file_name: str) -> list[dict]:
        return self._query(
            f"SELECT DISTINCT batch_id, file_name, MAX(uploaded_at) AS uploaded_at "
            f"FROM {table} WHERE file_name = %s GROUP BY batch_id, file_name "
            f"ORDER BY MAX(uploaded_at) DESC",
            (file_name,),
        )

    def raw_rows(self, table: str, batch_id: str) -> list[dict]:
        return self._query(
            f"SELECT id, row_number, raw_json FROM {table} WHERE batch_id = %s ORDER BY row_number",
            (batch_id,),
        )

    def etl_rows_by_batch(self, table: str, batch_id: str) -> list[dict]:
        return self._query(
            f"SELECT * FROM {table} WHERE batch_id = %s ORDER BY id",
            (batch_id,),
        )

    def etl_row_by_raw_id(self, table: str, raw_id: int) -> dict | None:
        rows = self._query(
            f"SELECT * FROM {table} WHERE raw_id = %s LIMIT 1",
            (raw_id,),
        )
        return rows[0] if rows else None

    def count_raw(self, table: str, batch_id: str) -> int:
        rows = self._query(
            f"SELECT COUNT(*) AS cnt FROM {table} WHERE batch_id = %s",
            (batch_id,),
        )
        return rows[0]["cnt"] if rows else 0

    def count_etl(self, table: str, batch_id: str) -> int:
        rows = self._query(
            f"SELECT COUNT(*) AS cnt FROM {table} WHERE batch_id = %s",
            (batch_id,),
        )
        return rows[0]["cnt"] if rows else 0


# ─────────────────────────────────────────────────────────────────────────────
# RULE 1 — File Metadata
# ─────────────────────────────────────────────────────────────────────────────

def validate_file_metadata(file_path: Path, file_type: str) -> RuleResult:
    """
    Rule 1: Capture filename, sheet names, row count, column count.
    Fails if file cannot be read or is empty.
    """
    issues: list[dict] = []
    details: dict = {}

    if not file_path.exists():
        return RuleResult(
            rule_id="R1",
            rule_name="File Metadata",
            file_type=file_type,
            passed=False,
            summary=f"File not found: {file_path}",
        )

    xl = pd.ExcelFile(str(file_path))
    sheet_info = {}
    for sheet in xl.sheet_names:
        df_full = pd.read_excel(xl, sheet_name=sheet, header=None, dtype=str)
        non_empty = df_full.dropna(how="all")
        sheet_info[sheet] = {
            "total_rows_including_header": len(df_full),
            "non_empty_rows": len(non_empty),
            "columns": len(df_full.columns),
        }

    details = {
        "file": file_path.name,
        "size_bytes": file_path.stat().st_size,
        "modified": str(datetime.fromtimestamp(file_path.stat().st_mtime).date()),
        "sheets": sheet_info,
    }

    # Pass if at least one sheet has meaningful data (>1 row).
    # Auxiliary sheets (Help, Summary) may legitimately be small.
    passed = any(v["non_empty_rows"] > 1 for v in sheet_info.values())
    if not passed:
        issues.append({"problem": "No sheet has more than 1 non-empty row — file appears empty"})

    summary = (
        f"File OK — {len(xl.sheet_names)} sheet(s), "
        + ", ".join(f"{s}: {v['non_empty_rows']} rows" for s, v in sheet_info.items())
    )
    return RuleResult(
        rule_id="R1", rule_name="File Metadata", file_type=file_type,
        passed=passed, summary=summary, details=details, issues=issues,
    )


# ─────────────────────────────────────────────────────────────────────────────
# RULE 2 — Column Name Validation
# ─────────────────────────────────────────────────────────────────────────────

def _detect_col_issues(col: str) -> list[str]:
    problems = []
    if col != col.lstrip():
        problems.append("leading_space")
    if col != col.rstrip():
        problems.append("trailing_space")
    for ch in col:
        if unicodedata.category(ch) in ("Cc", "Cf", "Zs") and ch not in (" ", "\t", "\n"):
            problems.append("hidden_char")
            break
    if any(ord(ch) > 127 for ch in col):
        problems.append("non_ascii")
    return problems


def validate_column_names(
    excel_result: ExcelReadResult,
    col_map: dict[str, str],
    file_type: str,
) -> RuleResult:
    """
    Rule 2: Validate column names for whitespace, duplicates, unicode, hidden chars.
    Report which columns are mapped vs unmapped.
    """
    issues = []
    raw_cols = excel_result.raw_columns

    # Check for duplicates
    seen: dict[str, int] = {}
    for c in raw_cols:
        seen[c] = seen.get(c, 0) + 1
    duplicates = [c for c, n in seen.items() if n > 1]
    if duplicates:
        issues.append({"problem": "duplicate_columns", "columns": duplicates})

    # Check for character issues (on the original pre-cleaned column names)
    xl = pd.ExcelFile(excel_result.file_path)
    raw_df = pd.read_excel(
        xl, sheet_name=excel_result.sheet,
        header=excel_result.header_row, dtype=str, nrows=0,
    )
    original_cols = [str(c) for c in raw_df.columns]
    char_issues = []
    for orig in original_cols:
        problems = _detect_col_issues(orig)
        if problems:
            char_issues.append({"column": orig, "problems": problems})
    if char_issues:
        issues.append({"problem": "column_char_issues", "columns": char_issues})

    # Unmapped columns
    if excel_result.unmapped_columns:
        issues.append({
            "problem": "unmapped_columns",
            "count": len(excel_result.unmapped_columns),
            "columns": excel_result.unmapped_columns,
        })

    # Build full mapping report
    full_map = {}
    for raw_col in raw_cols:
        canonical = excel_result.mapped_columns.get(raw_col, "<unmapped>")
        full_map[raw_col] = canonical

    details = {
        "original_excel_columns": original_cols,
        "cleaned_columns": raw_cols,
        "column_map": full_map,
        "duplicates": duplicates,
        "char_issues": char_issues,
        "unmapped_count": len(excel_result.unmapped_columns),
        "unmapped": excel_result.unmapped_columns,
        "sub_header_dropped": excel_result.sub_header_dropped,
    }

    # Only fail R2 for chars that the ingestor CANNOT silently fix:
    # hidden control chars and non-ASCII. Leading/trailing spaces are stripped
    # by _clean_col (same as ingestor._clean_col_name) so they don't cause data loss.
    breaking_char_issues = [
        ci for ci in char_issues
        if set(ci["problems"]) - {"leading_space", "trailing_space"}
    ]
    passed = not duplicates and not breaking_char_issues
    summary = (
        f"{len(raw_cols)} cols — "
        f"{len(excel_result.mapped_columns)} mapped, "
        f"{len(excel_result.unmapped_columns)} unmapped"
        + (f", {len(duplicates)} duplicates" if duplicates else "")
        + (f", {len(char_issues)} char-issues" if char_issues else "")
    )
    return RuleResult(
        rule_id="R2", rule_name="Column Names", file_type=file_type,
        passed=passed, summary=summary, details=details, issues=issues,
    )


# ─────────────────────────────────────────────────────────────────────────────
# RULE 3 — Row Count: Excel vs raw table
# ─────────────────────────────────────────────────────────────────────────────

def validate_row_count(
    excel_result: ExcelReadResult,
    db: DbClient,
    raw_table: str,
    batch_id: str,
    file_type: str,
) -> RuleResult:
    """
    Rule 3: Excel data-row count must equal raw table row count for this batch.
    Fails on any mismatch.
    """
    excel_count = len(excel_result.df)
    db_count = db.count_raw(raw_table, batch_id)
    match = excel_count == db_count

    details = {
        "excel_data_rows": excel_count,
        "raw_table_rows": db_count,
        "delta": db_count - excel_count,
        "batch_id": str(batch_id),
        "table": raw_table,
        "sub_header_was_dropped": excel_result.sub_header_dropped,
        "empty_rows_dropped": excel_result.empty_rows_dropped,
    }

    issues = []
    if not match:
        issues.append({
            "problem": "row_count_mismatch",
            "excel": excel_count,
            "db": db_count,
            "delta": db_count - excel_count,
        })

    summary = (
        f"Excel={excel_count}, DB={db_count}"
        + (" [MATCH]" if match else f" [MISMATCH] delta={db_count - excel_count}")
    )
    return RuleResult(
        rule_id="R3", rule_name="Row Count", file_type=file_type,
        passed=match, summary=summary, details=details, issues=issues,
    )


# ─────────────────────────────────────────────────────────────────────────────
# RULE 4 — Per-row Integrity: find every Excel row in the raw table
# ─────────────────────────────────────────────────────────────────────────────

def validate_row_integrity(
    excel_result: ExcelReadResult,
    db: DbClient,
    raw_table: str,
    batch_id: str,
    file_type: str,
    key_col: str = "order_item_id",
) -> RuleResult:
    """
    Rule 4: For every Excel row find its match in raw_json by order_item_id.
    Reports rows missing from DB and unexpected extra rows.
    """
    df = excel_result.df

    if key_col not in df.columns:
        return RuleResult(
            rule_id="R4", rule_name="Row Integrity", file_type=file_type,
            passed=False,
            summary=f"Key column '{key_col}' not found in DataFrame columns: {list(df.columns)}",
        )

    # Excel keys — apply the same OI: stripping as transformer does for orders
    def _normalise_key(v) -> str | None:
        s = _clean_val(v)
        if s is None:
            return None
        if file_type == "orders" and s.startswith("OI:"):
            s = s[3:]
        return s

    excel_keys_raw = [_normalise_key(v) for v in df[key_col]]
    excel_keys = [k for k in excel_keys_raw if k is not None]
    excel_key_set = set(excel_keys)

    # DB keys
    raw_rows = db.raw_rows(raw_table, batch_id)
    db_keys = set()
    for row in raw_rows:
        k = row["raw_json"].get(key_col)
        if k:
            if file_type == "orders" and str(k).startswith("OI:"):
                k = str(k)[3:]
            db_keys.add(str(k))

    missing_in_db = sorted(excel_key_set - db_keys)
    extra_in_db = sorted(db_keys - excel_key_set)

    # Duplicates in Excel
    from collections import Counter
    key_counts = Counter(excel_keys)
    excel_duplicates = {k: v for k, v in key_counts.items() if v > 1}

    issues = []
    if missing_in_db:
        issues.append({"problem": "rows_missing_in_db", "count": len(missing_in_db), "keys": missing_in_db[:50]})
    if extra_in_db:
        issues.append({"problem": "extra_rows_in_db", "count": len(extra_in_db), "keys": extra_in_db[:50]})

    null_keys_count = sum(1 for k in excel_keys_raw if k is None)
    if null_keys_count:
        issues.append({"problem": "null_order_item_ids_in_excel", "count": null_keys_count})

    passed = not missing_in_db and not extra_in_db
    details = {
        "excel_key_count": len(excel_key_set),
        "db_key_count": len(db_keys),
        "missing_in_db": len(missing_in_db),
        "extra_in_db": len(extra_in_db),
        "null_keys_in_excel": null_keys_count,
        "duplicate_keys_in_excel": len(excel_duplicates),
        "duplicates_sample": dict(list(excel_duplicates.items())[:10]),
    }
    summary = (
        f"Keys: Excel={len(excel_key_set)}, DB={len(db_keys)} — "
        f"missing={len(missing_in_db)}, extra={len(extra_in_db)}"
    )
    return RuleResult(
        rule_id="R4", rule_name="Row Integrity", file_type=file_type,
        passed=passed, summary=summary, details=details, issues=issues,
    )


# ─────────────────────────────────────────────────────────────────────────────
# RULE 5 — Column Values: Excel vs raw_json
# ─────────────────────────────────────────────────────────────────────────────

def validate_column_values(
    excel_result: ExcelReadResult,
    db: DbClient,
    raw_table: str,
    batch_id: str,
    file_type: str,
    sample_size: int | None = None,
) -> RuleResult:
    """
    Rule 5: For every column, compare Excel value against raw_json value.
    Applies the same _clean_val transformation to Excel values before comparison.

    When multiple Excel columns map to the same canonical (e.g. both
    order_approval_date and dispatched_date → dispatch_date), the ingestor
    stores the last-written value. We identify the "winner" (the source whose
    Excel value matches what raw_json stores) and skip the others to avoid
    false positives from known ingestor column-merge behaviour.
    """
    from collections import defaultdict

    df = excel_result.df_raw   # raw column names (before rename)
    col_map = excel_result.mapped_columns

    raw_rows = db.raw_rows(raw_table, batch_id)
    if len(raw_rows) == 0:
        return RuleResult(
            rule_id="R5", rule_name="Column Values", file_type=file_type,
            passed=False, summary="No rows found in raw table for this batch",
        )

    # Build lookup: row_number → raw_json
    db_by_row: dict[int, dict] = {r["row_number"]: r["raw_json"] for r in raw_rows}

    # Identify canonicals that have multiple source raw columns (collision candidates)
    canonical_to_raws: dict[str, list[str]] = defaultdict(list)
    for rc, cn in col_map.items():
        canonical_to_raws[cn].append(rc)
    multi_source = {cn for cn, rcs in canonical_to_raws.items() if len(rcs) > 1}

    # Apply sample
    indices = list(range(len(df)))
    if sample_size and sample_size < len(indices):
        import random
        random.seed(42)
        indices = sorted(random.sample(indices, sample_size))

    mismatches: list[dict] = []
    checked = 0

    for idx in indices:
        row = df.iloc[idx]
        # row_number in DB is 1-based index after header skip and sub-header skip
        row_number = idx + 1
        db_json = db_by_row.get(row_number)
        if db_json is None:
            mismatches.append({
                "excel_row": row_number,
                "problem": "row_not_found_in_db",
            })
            continue

        # For multi-source canonicals, determine which raw_col "won"
        # (the one whose cleaned Excel value matches raw_json's canonical value).
        canonical_winner: dict[str, str] = {}
        for cn in multi_source:
            db_cn_val = db_json.get(cn)
            for rc in canonical_to_raws[cn]:
                if _clean_val(row.get(rc)) == db_cn_val:
                    canonical_winner[cn] = rc
                    break

        for raw_col in excel_result.raw_columns:
            canonical = col_map.get(raw_col, raw_col)

            # Skip secondary sources for multi-source canonicals to avoid false
            # mismatches from intentional ingestor column-merge behaviour.
            if canonical in multi_source and canonical_winner.get(canonical) != raw_col:
                continue

            excel_val = row.get(raw_col)
            expected_val = _clean_val(excel_val)

            db_val = db_json.get(canonical)

            # Also check under the raw (unmapped) column name
            if db_val is None and canonical != raw_col:
                db_val = db_json.get(raw_col)

            if expected_val != db_val:
                mismatches.append({
                    "excel_row": row_number,
                    "excel_col": raw_col,
                    "canonical_col": canonical,
                    "excel_value": str(excel_val)[:100],
                    "expected_in_db": str(expected_val)[:100],
                    "actual_in_db": str(db_val)[:100] if db_val is not None else None,
                })
        checked += 1

    issues = []
    if mismatches:
        issues.append({
            "problem": "value_mismatches",
            "count": len(mismatches),
            "sample": mismatches[:30],
        })

    passed = len(mismatches) == 0
    details = {
        "rows_checked": checked,
        "total_comparisons": checked * len(excel_result.raw_columns),
        "mismatches": len(mismatches),
    }
    summary = (
        f"Checked {checked} rows × {len(excel_result.raw_columns)} cols — "
        f"{len(mismatches)} value mismatch(es)"
    )
    return RuleResult(
        rule_id="R5", rule_name="Column Values", file_type=file_type,
        passed=passed, summary=summary, details=details, issues=issues,
    )


# ─────────────────────────────────────────────────────────────────────────────
# RULE 6 — NULL Integrity
# ─────────────────────────────────────────────────────────────────────────────

def validate_null_integrity(
    excel_result: ExcelReadResult,
    db: DbClient,
    raw_table: str,
    etl_table: str,
    batch_id: str,
    file_type: str,
    non_null_cols: list[str],        # columns that must never be None in ETL
    nullable_cols: list[str],        # columns expected to be sometimes None
) -> RuleResult:
    """
    Rule 6: Detect unexpected NULLs introduced at each pipeline stage.
    Checks:
      A) Non-null Excel values → None in raw_json (unexpected NULL injection)
      B) Required ETL columns (non_null_cols) that are None (NULL in ETL)
    """
    issues = []

    # A) Check Excel non-null → raw_json None
    from collections import defaultdict

    df_raw = excel_result.df_raw
    col_map = excel_result.mapped_columns
    raw_rows = db.raw_rows(raw_table, batch_id)
    db_by_row: dict[int, dict] = {r["row_number"]: r["raw_json"] for r in raw_rows}

    # Multi-source canonical detection (same as R5)
    canonical_to_raws_r6: dict[str, list[str]] = defaultdict(list)
    for rc, cn in col_map.items():
        canonical_to_raws_r6[cn].append(rc)
    multi_source_r6 = {cn for cn, rcs in canonical_to_raws_r6.items() if len(rcs) > 1}

    null_injections: list[dict] = []
    for idx in range(len(df_raw)):
        row = df_raw.iloc[idx]
        row_number = idx + 1
        db_json = db_by_row.get(row_number, {})

        # Determine canonical winners for multi-source cols
        canonical_winner_r6: dict[str, str] = {}
        for cn in multi_source_r6:
            db_cn_val = db_json.get(cn)
            for rc in canonical_to_raws_r6[cn]:
                if _clean_val(row.get(rc)) == db_cn_val:
                    canonical_winner_r6[cn] = rc
                    break

        for raw_col in excel_result.raw_columns:
            # Skip secondary sources for multi-source canonicals
            canonical = col_map.get(raw_col, raw_col)
            if canonical in multi_source_r6 and canonical_winner_r6.get(canonical) != raw_col:
                continue

            excel_val = row.get(raw_col)
            # If Excel has a real value but raw_json has None
            is_excel_non_null = not (
                excel_val is None
                or (isinstance(excel_val, float) and pd.isna(excel_val))
                or str(excel_val).strip() in ("", "nan", "NaN", "None")
            )
            if is_excel_non_null:
                db_val = db_json.get(canonical) or db_json.get(raw_col)
                if db_val is None:
                    null_injections.append({
                        "excel_row": row_number,
                        "column": raw_col,
                        "canonical": canonical,
                        "excel_value": str(excel_val)[:80],
                    })

    if null_injections:
        issues.append({
            "problem": "unexpected_null_in_raw",
            "count": len(null_injections),
            "sample": null_injections[:20],
        })

    # B) Check required ETL columns for NULLs
    etl_rows = db.etl_rows_by_batch(etl_table, batch_id)
    etl_null_violations: list[dict] = []
    for row in etl_rows:
        for col in non_null_cols:
            if col in row and row[col] is None:
                etl_null_violations.append({
                    "etl_id": row.get("id"),
                    "order_item_id": row.get("order_item_id"),
                    "column": col,
                })

    if etl_null_violations:
        issues.append({
            "problem": "required_etl_column_is_null",
            "count": len(etl_null_violations),
            "sample": etl_null_violations[:20],
        })

    # C) Check ETL rows with validation_errors
    invalid_etl = [r for r in etl_rows if r.get("is_valid") is False]
    if invalid_etl:
        issues.append({
            "problem": "invalid_etl_rows",
            "count": len(invalid_etl),
            "sample": [
                {"id": r["id"], "order_item_id": r.get("order_item_id"), "errors": r.get("validation_errors")}
                for r in invalid_etl[:20]
            ],
        })

    passed = len(null_injections) == 0 and len(etl_null_violations) == 0
    details = {
        "null_injections_raw": len(null_injections),
        "etl_null_violations": len(etl_null_violations),
        "invalid_etl_rows": len(invalid_etl),
        "etl_total_rows": len(etl_rows),
    }
    summary = (
        f"NULL injections (raw)={len(null_injections)}, "
        f"ETL required-col NULLs={len(etl_null_violations)}, "
        f"invalid ETL rows={len(invalid_etl)}"
    )
    return RuleResult(
        rule_id="R6", rule_name="NULL Integrity", file_type=file_type,
        passed=passed, summary=summary, details=details, issues=issues,
    )


# ─────────────────────────────────────────────────────────────────────────────
# RULE 7 — Numeric Precision
# ─────────────────────────────────────────────────────────────────────────────

def validate_numeric_precision(
    db: DbClient,
    raw_table: str,
    etl_table: str,
    batch_id: str,
    file_type: str,
    numeric_cols: list[str],
    optional_cols: list[str] | None = None,
) -> RuleResult:
    """
    Rule 7: For every numeric ETL column, independently apply parse_decimal
    to the raw_json value and compare against the stored ETL Decimal.
    Detects any rounding loss or silent zeroing.
    """
    optional_cols = optional_cols or []
    raw_rows = db.raw_rows(raw_table, batch_id)
    db_by_raw_id = {r["id"]: r["raw_json"] for r in raw_rows}

    etl_rows = db.etl_rows_by_batch(etl_table, batch_id)

    _TWO_DP = Decimal("0.01")
    mismatches: list[dict] = []      # value differs even after 2dp rounding
    rounding_deltas: list[dict] = [] # raw had >2dp — DB rounded as expected
    silent_zeros: list[dict] = []    # raw had non-zero value but ETL stored 0

    for etl_row in etl_rows:
        raw_id = etl_row.get("raw_id")
        raw_json = db_by_raw_id.get(raw_id, {})

        for col in numeric_cols:
            raw_val = raw_json.get(col)
            etl_val = etl_row.get(col)

            if col in optional_cols:
                expected = _parse_decimal_optional(raw_val)
            else:
                expected = _parse_decimal(raw_val)

            if etl_val is None:
                actual = None
            else:
                actual = Decimal(str(etl_val))

            # PostgreSQL Numeric(15,2) rounds to 2dp on insert using ROUND_HALF_UP.
            # Apply the same rounding so Python's default ROUND_HALF_EVEN doesn't
            # produce false positives for values ending in exactly .5 at the 3rd dp.
            if expected is not None:
                expected_2dp = expected.quantize(_TWO_DP, rounding=ROUND_HALF_UP)
            else:
                expected_2dp = None

            if expected_2dp != actual:
                mismatches.append({
                    "etl_id": etl_row.get("id"),
                    "order_item_id": etl_row.get("order_item_id"),
                    "column": col,
                    "raw_json_value": str(raw_val),
                    "expected_decimal": str(expected),
                    "expected_after_2dp": str(expected_2dp),
                    "actual_etl_value": str(actual),
                })
            elif expected is not None and expected != expected_2dp:
                # Same after rounding, but raw had >2dp — log for info
                rounding_deltas.append({
                    "etl_id": etl_row.get("id"),
                    "column": col,
                    "raw_value": str(raw_val),
                    "raw_full_precision": str(expected),
                    "stored_rounded": str(actual),
                    "delta": str(abs(expected - actual)),
                })

            # Detect silent zeroing: raw has a non-trivial numeric data but ETL
            # stored 0. Use expected_2dp so sub-paisa values (e.g. -0.001) that
            # round to 0.00 are not false positives.
            if raw_val and raw_val not in ("None", "nan", "0", "0.0", "0.00"):
                if actual == Decimal("0") and expected_2dp is not None and expected_2dp != Decimal("0"):
                    silent_zeros.append({
                        "etl_id": etl_row.get("id"),
                        "order_item_id": etl_row.get("order_item_id"),
                        "column": col,
                        "raw_value": raw_val,
                        "expected_2dp": str(expected_2dp),
                    })

    # Also verify precision: check that no raw value has more decimal places than DB allows (15,2)
    precision_loss: list[dict] = []
    for raw_row in raw_rows:
        raw_json = raw_row["raw_json"]
        for col in numeric_cols:
            raw_val = raw_json.get(col)
            if raw_val is None:
                continue
            cleaned = str(raw_val).replace(",", "").replace("₹", "").strip()
            try:
                d = Decimal(cleaned)
                if abs(d.as_tuple().exponent) > 2:
                    # More than 2 decimal places — check if DB would round
                    db_d = d.quantize(Decimal("0.01"))
                    if db_d != d:
                        precision_loss.append({
                            "row_number": raw_row["row_number"],
                            "column": col,
                            "raw_value": raw_val,
                            "after_2dp_rounding": str(db_d),
                        })
            except InvalidOperation:
                pass

    issues = []
    if mismatches:
        issues.append({"problem": "numeric_mismatch", "count": len(mismatches), "sample": mismatches[:20]})
    if silent_zeros:
        issues.append({"problem": "silent_zero", "count": len(silent_zeros), "sample": silent_zeros[:20]})
    if precision_loss:
        issues.append({
            "problem": "precision_schema_note",
            "note": "Raw has >2dp; DB column is Numeric(15,2) -- sub-paisa precision lost on insert",
            "count": len(precision_loss), "sample": precision_loss[:10],
        })
    if rounding_deltas:
        issues.append({
            "problem": "rounding_delta_info",
            "note": "Values match after Numeric(15,2) rounding -- not a failure",
            "count": len(rounding_deltas), "sample": rounding_deltas[:10],
        })

    passed = len(mismatches) == 0 and len(silent_zeros) == 0
    details = {
        "rows_checked": len(etl_rows),
        "columns_checked": numeric_cols,
        "mismatches": len(mismatches),
        "silent_zeros": len(silent_zeros),
        "rounding_deltas_info": len(rounding_deltas),
        "precision_schema_note": len(precision_loss),
    }
    summary = (
        f"Checked {len(etl_rows)} rows x {len(numeric_cols)} cols -- "
        f"mismatches={len(mismatches)}, silent_zeros={len(silent_zeros)}, "
        f"rounding_info={len(rounding_deltas)}"
    )
    return RuleResult(
        rule_id="R7", rule_name="Numeric Precision", file_type=file_type,
        passed=passed, summary=summary, details=details, issues=issues,
    )


# ─────────────────────────────────────────────────────────────────────────────
# RULE 8 — Date Integrity
# ─────────────────────────────────────────────────────────────────────────────

def validate_date_integrity(
    db: DbClient,
    raw_table: str,
    etl_table: str,
    batch_id: str,
    file_type: str,
    date_cols: list[str],
) -> RuleResult:
    """
    Rule 8: Independently parse date strings from raw_json and compare against
    ETL date columns. Detects unparseable formats and timezone shifts.
    """
    raw_rows = db.raw_rows(raw_table, batch_id)
    db_by_raw_id = {r["id"]: r["raw_json"] for r in raw_rows}

    etl_rows = db.etl_rows_by_batch(etl_table, batch_id)

    mismatches: list[dict] = []
    parse_failures: list[dict] = []   # raw had a date string but ETL got None
    format_inventory: dict[str, int] = {}   # date format → occurrence count

    for etl_row in etl_rows:
        raw_id = etl_row.get("raw_id")
        raw_json = db_by_raw_id.get(raw_id, {})

        for col in date_cols:
            raw_val = raw_json.get(col)
            etl_val = etl_row.get(col)

            expected = _parse_date(raw_val)

            # Track format for inventory
            if raw_val:
                _fmt = _detect_date_format(str(raw_val))
                format_inventory[_fmt] = format_inventory.get(_fmt, 0) + 1

            # ETL stores as Python date
            if etl_val is None:
                actual = None
            elif isinstance(etl_val, datetime):
                actual = etl_val.date()
            elif isinstance(etl_val, date):
                actual = etl_val
            else:
                actual = _parse_date(str(etl_val))

            if expected != actual:
                mismatches.append({
                    "etl_id": etl_row.get("id"),
                    "order_item_id": etl_row.get("order_item_id"),
                    "column": col,
                    "raw_value": str(raw_val),
                    "expected": str(expected),
                    "actual_etl": str(actual),
                })

            # Detect parse failures: raw had a non-null date but ETL got None
            if raw_val and raw_val not in ("None", "nan", "NaT") and expected is None and actual is None:
                parse_failures.append({
                    "etl_id": etl_row.get("id"),
                    "order_item_id": etl_row.get("order_item_id"),
                    "column": col,
                    "raw_value": str(raw_val),
                })

    issues = []
    if mismatches:
        issues.append({"problem": "date_mismatch", "count": len(mismatches), "sample": mismatches[:20]})
    if parse_failures:
        issues.append({"problem": "date_parse_failure", "count": len(parse_failures), "sample": parse_failures[:20]})

    # Check for timezone confusion: datetime → date conversion should drop time but not shift day
    tz_suspect: list[dict] = []
    for raw_row in raw_rows:
        for col in date_cols:
            rv = raw_row["raw_json"].get(col)
            if rv and "00:00:00" not in str(rv):  # has time component
                # If time component is non-zero, conversion could shift the date
                try:
                    dt = datetime.fromisoformat(str(rv).replace(" ", "T"))
                    if dt.hour != 0 or dt.minute != 0:
                        tz_suspect.append({
                            "row_number": raw_row["row_number"],
                            "column": col,
                            "raw_value": rv,
                        })
                except (ValueError, TypeError):
                    pass

    if tz_suspect:
        issues.append({"problem": "non_midnight_time_in_date", "count": len(tz_suspect), "sample": tz_suspect[:10]})

    passed = len(mismatches) == 0 and len(parse_failures) == 0
    details = {
        "rows_checked": len(etl_rows),
        "date_columns": date_cols,
        "mismatches": len(mismatches),
        "parse_failures": len(parse_failures),
        "tz_suspects": len(tz_suspect),
        "date_format_inventory": format_inventory,
    }
    summary = (
        f"Checked {len(etl_rows)} rows × {len(date_cols)} cols — "
        f"mismatches={len(mismatches)}, parse_failures={len(parse_failures)}"
    )
    return RuleResult(
        rule_id="R8", rule_name="Date Integrity", file_type=file_type,
        passed=passed, summary=summary, details=details, issues=issues,
    )


def _detect_date_format(s: str) -> str:
    s = s.strip()
    if re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", s):
        return "YYYY-MM-DD HH:MM:SS"
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return "YYYY-MM-DD"
    if re.match(r"^\d{2}-\d{2}-\d{4}$", s):
        return "DD-MM-YYYY"
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        return "DD/MM/YYYY"
    if re.match(r"^\d{2} [A-Za-z]+ \d{4}$", s):
        return "DD Mon YYYY"
    return f"unknown({s[:20]})"


# ─────────────────────────────────────────────────────────────────────────────
# Report Generation
# ─────────────────────────────────────────────────────────────────────────────

def generate_reports(report: AuditReport, output_dir: Path) -> list[Path]:
    """
    Generate 5 Excel reports from the AuditReport.
    Returns list of file paths created.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed — skipping report generation")
        return []

    output_dir.mkdir(parents=True, exist_ok=True)
    created = []

    _HDR = PatternFill("solid", fgColor="1F3864")
    _HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    _PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
    _FAIL_FILL = PatternFill("solid", fgColor="FFCCCC")
    _WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
    _TITLE_FONT = Font(bold=True, size=13, color="1F3864")

    def _hdr(ws, row: int, cols: list[str]) -> None:
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill = _HDR
            cell.font = _HDR_FONT
            cell.alignment = Alignment(horizontal="center")

    def _title(ws, text: str) -> None:
        ws["A1"] = text
        ws["A1"].font = _TITLE_FONT
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(italic=True, size=9, color="666666")

    # ── Report 1: Data Integrity Report ──────────────────────────────────────
    wb1 = Workbook()
    ws = wb1.active
    ws.title = "Audit Summary"
    _title(ws, "Data Integrity Audit — Summary")
    _hdr(ws, 4, ["Rule ID", "Rule Name", "File Type", "Status", "Summary", "Issue Count"])
    for i, r in enumerate(report.results, 5):
        fill = _PASS_FILL if r.passed else _FAIL_FILL
        ws.cell(i, 1, r.rule_id)
        ws.cell(i, 2, r.rule_name)
        ws.cell(i, 3, r.file_type)
        cell = ws.cell(i, 4, r.status)
        cell.fill = fill
        cell.font = Font(bold=True)
        ws.cell(i, 5, r.summary)
        ws.cell(i, 6, len(r.issues))
    for col, w in [(1, 8), (2, 22), (3, 14), (4, 8), (5, 80), (6, 12)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    # Details sheet per rule
    for r in report.results:
        detail_ws = wb1.create_sheet(f"{r.rule_id}_{r.file_type[:3]}")
        detail_ws["A1"] = f"{r.rule_id}: {r.rule_name} — {r.file_type}"
        detail_ws["A1"].font = _TITLE_FONT
        detail_ws["A2"] = r.summary
        detail_ws["A3"] = f"PASS: {r.passed}"
        detail_ws["A3"].font = Font(bold=True, color="006100" if r.passed else "CC0000")
        # Write details as key-value
        row = 5
        detail_ws.cell(row, 1, "Details").font = Font(bold=True)
        row += 1
        for k, v in r.details.items():
            detail_ws.cell(row, 1, str(k))
            detail_ws.cell(row, 2, str(v)[:200])
            row += 1
        if r.issues:
            row += 1
            detail_ws.cell(row, 1, "Issues").font = Font(bold=True)
            row += 1
            for issue in r.issues:
                detail_ws.cell(row, 1, json.dumps(issue)[:500])
                row += 1
        detail_ws.column_dimensions["A"].width = 30
        detail_ws.column_dimensions["B"].width = 80

    p1 = output_dir / "01_data_integrity_report.xlsx"
    wb1.save(str(p1))
    created.append(p1)

    # ── Report 2: Missing Rows Report ─────────────────────────────────────────
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Missing Rows"
    _title(ws2, "Missing Rows Report")
    row = 4
    missing_results = [r for r in report.results if r.rule_id == "R4"]
    if missing_results:
        _hdr(ws2, row, ["File Type", "Missing in DB", "Extra in DB", "Null Keys", "Detail"])
        row += 1
        for r in missing_results:
            ws2.cell(row, 1, r.file_type)
            ws2.cell(row, 2, r.details.get("missing_in_db", 0))
            ws2.cell(row, 3, r.details.get("extra_in_db", 0))
            ws2.cell(row, 4, r.details.get("null_keys_in_excel", 0))
            ws2.cell(row, 5, r.summary)
            if r.issues:
                for iss in r.issues:
                    if "keys" in iss:
                        row += 1
                        ws2.cell(row, 1, iss.get("problem", ""))
                        ws2.cell(row, 2, iss.get("count", 0))
                        for j, k in enumerate(iss["keys"][:100], 3):
                            ws2.cell(row, j, str(k))
            row += 2
    for col, w in [(1, 16), (2, 16), (3, 16), (4, 14), (5, 60)]:
        ws2.column_dimensions[get_column_letter(col)].width = w
    p2 = output_dir / "02_missing_rows_report.xlsx"
    wb2.save(str(p2))
    created.append(p2)

    # ── Report 3: Column Mapping Report ──────────────────────────────────────
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Column Mapping"
    _title(ws3, "Column Mapping Report")
    row = 4
    for r in report.results:
        if r.rule_id != "R2":
            continue
        ws3.cell(row, 1, f"File: {r.file_type}").font = Font(bold=True)
        row += 1
        _hdr(ws3, row, ["Excel Column (cleaned)", "Canonical DB Column", "Status"])
        row += 1
        col_map = r.details.get("column_map", {})
        for exc, can in col_map.items():
            ws3.cell(row, 1, exc)
            ws3.cell(row, 2, can)
            status_cell = ws3.cell(row, 3, "UNMAPPED" if can == "<unmapped>" else "MAPPED")
            status_cell.fill = _WARN_FILL if can == "<unmapped>" else _PASS_FILL
            row += 1
        row += 2
    for col, w in [(1, 50), (2, 40), (3, 12)]:
        ws3.column_dimensions[get_column_letter(col)].width = w
    p3 = output_dir / "03_column_mapping_report.xlsx"
    wb3.save(str(p3))
    created.append(p3)

    # ── Report 4: Numeric Precision Report ───────────────────────────────────
    wb4 = Workbook()
    ws4 = wb4.active
    ws4.title = "Numeric Precision"
    _title(ws4, "Numeric Precision Report")
    row = 4
    _hdr(ws4, row, ["File Type", "Rows Checked", "Mismatches", "Silent Zeros", "Precision Warnings", "Status"])
    row += 1
    for r in report.results:
        if r.rule_id != "R7":
            continue
        fill = _PASS_FILL if r.passed else _FAIL_FILL
        ws4.cell(row, 1, r.file_type)
        ws4.cell(row, 2, r.details.get("rows_checked", 0))
        ws4.cell(row, 3, r.details.get("mismatches", 0))
        ws4.cell(row, 4, r.details.get("silent_zeros", 0))
        ws4.cell(row, 5, r.details.get("precision_loss_candidates", 0))
        ws4.cell(row, 6, r.status).fill = fill
        row += 1
        if r.issues:
            for iss in r.issues:
                if "sample" in iss:
                    _hdr(ws4, row, ["Problem", "ETL ID", "Order Item ID", "Column", "Raw Value", "Expected", "Actual"])
                    row += 1
                    for s in iss["sample"]:
                        ws4.cell(row, 1, iss.get("problem", ""))
                        ws4.cell(row, 2, str(s.get("etl_id", "")))
                        ws4.cell(row, 3, str(s.get("order_item_id", "")))
                        ws4.cell(row, 4, str(s.get("column", "")))
                        ws4.cell(row, 5, str(s.get("raw_json_value", s.get("raw_value", "")))[:50])
                        ws4.cell(row, 6, str(s.get("expected_decimal", s.get("expected", ""))))
                        ws4.cell(row, 7, str(s.get("actual_etl_value", "")))
                        row += 1
            row += 1
    for col, w in [(1, 16), (2, 14), (3, 14), (4, 14), (5, 20), (6, 10), (7, 20)]:
        ws4.column_dimensions[get_column_letter(col)].width = w
    p4 = output_dir / "04_numeric_precision_report.xlsx"
    wb4.save(str(p4))
    created.append(p4)

    # ── Report 5: Date Conversion Report ─────────────────────────────────────
    wb5 = Workbook()
    ws5 = wb5.active
    ws5.title = "Date Integrity"
    _title(ws5, "Date Conversion Report")
    row = 4
    for r in report.results:
        if r.rule_id != "R8":
            continue
        ws5.cell(row, 1, f"File: {r.file_type}").font = Font(bold=True)
        row += 1
        ws5.cell(row, 1, f"Mismatches: {r.details.get('mismatches', 0)}")
        ws5.cell(row, 2, f"Parse Failures: {r.details.get('parse_failures', 0)}")
        ws5.cell(row, 3, f"TZ Suspects: {r.details.get('tz_suspects', 0)}")
        row += 1

        # Date format inventory
        fmt_inv = r.details.get("date_format_inventory", {})
        if fmt_inv:
            ws5.cell(row, 1, "Date Formats Found").font = Font(bold=True)
            row += 1
            _hdr(ws5, row, ["Format", "Count"])
            row += 1
            for fmt, cnt in sorted(fmt_inv.items(), key=lambda x: -x[1]):
                ws5.cell(row, 1, fmt)
                ws5.cell(row, 2, cnt)
                row += 1

        if r.issues:
            row += 1
            _hdr(ws5, row, ["Problem", "ETL ID", "Order Item ID", "Column", "Raw Value", "Expected", "Actual"])
            row += 1
            for iss in r.issues:
                for s in iss.get("sample", []):
                    ws5.cell(row, 1, iss.get("problem", ""))
                    ws5.cell(row, 2, str(s.get("etl_id", "")))
                    ws5.cell(row, 3, str(s.get("order_item_id", "")))
                    ws5.cell(row, 4, str(s.get("column", "")))
                    ws5.cell(row, 5, str(s.get("raw_value", ""))[:50])
                    ws5.cell(row, 6, str(s.get("expected", "")))
                    ws5.cell(row, 7, str(s.get("actual_etl", "")))
                    row += 1
        row += 2
    for col, w in [(1, 28), (2, 14), (3, 25), (4, 18), (5, 30), (6, 14), (7, 14)]:
        ws5.column_dimensions[get_column_letter(col)].width = w
    p5 = output_dir / "05_date_conversion_report.xlsx"
    wb5.save(str(p5))
    created.append(p5)

    return created
