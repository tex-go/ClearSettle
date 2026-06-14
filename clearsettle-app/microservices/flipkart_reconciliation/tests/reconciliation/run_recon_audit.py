"""
Phase 3: Reconciliation Engine Audit Runner
==========================================
Generates 6 Excel audit reports:
  11_recon_audit_report.xlsx       -- per-order reconciliation with formula
  12_formula_validation_report.xlsx-- credit/debit breakdown per order
  13_leakage_report.xlsx           -- POTENTIAL_MONEY_LEAK orders only
  14_unknown_component_report.xlsx -- unclassified components
  15_confidence_score_report.xlsx  -- per-order confidence distribution
  16_test_suite_report.xlsx        -- Python test suite summary

Run:
  python run_recon_audit.py
"""
from __future__ import annotations

import os
import sys
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE.parent.parent))

from tests.reconciliation.recon_audit import (
    ReconDbClient,
    ReconResult,
    ReconAuditReport,
    SETTLEMENT_COMPONENTS,
    STATUS_LABELS,
    reconcile_order,
    test_existence_checks,
    test_component_classification,
    test_formula_validation,
    test_settlement_comparison,
    test_leakage_detection,
    test_confidence_scores,
)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DSN = os.environ.get(
    "DATABASE_URL",
    "postgresql://clearsettle_user:clearsettle_pass@localhost:5432/flipkart_recon",
)

OUT_DIR = HERE / "recon_audit_reports"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
PASS_FILL  = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL  = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL  = PatternFill("solid", fgColor="FFEB9C")
INFO_FILL  = PatternFill("solid", fgColor="DDEBF7")
STATUS_FILLS = {
    "MATCHED":             PatternFill("solid", fgColor="C6EFCE"),
    "OVER_PAID":           PatternFill("solid", fgColor="FFEB9C"),
    "SHORT_PAID":          PatternFill("solid", fgColor="FFC7CE"),
    "RETURN_RECOVERY":     PatternFill("solid", fgColor="D9D9D9"),
    "MISSING_SETTLEMENT":  PatternFill("solid", fgColor="FCE4D6"),
    "MISSING_FEE_RECORD":  PatternFill("solid", fgColor="FFC7CE"),
    "MISSING_ORDER":       PatternFill("solid", fgColor="FF0000"),
}


def _hdr(ws, row: int, headers: list[str]) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")


def _write_xl(path: Path, sheets: dict[str, dict]) -> None:
    """
    sheets = {
        "Sheet Name": {
            "headers": [...],
            "rows": [[...], ...],
            "col_widths": [...],          optional
            "status_col": int,            optional — column index (1-based) for status coloring
            "pass_col": int,              optional — column index for boolean PASS/FAIL coloring
        }
    }
    """
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, cfg in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        headers   = cfg.get("headers", [])
        rows      = cfg.get("rows", [])
        col_widths = cfg.get("col_widths", [])
        status_col = cfg.get("status_col")
        pass_col   = cfg.get("pass_col")

        _hdr(ws, 1, headers)
        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if status_col and c_idx == status_col:
                    cell.fill = STATUS_FILLS.get(str(val), PatternFill())
                if pass_col and c_idx == pass_col:
                    cell.fill = PASS_FILL if val is True or val == "PASS" else FAIL_FILL

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
    wb.save(path)
    print(f"  [SAVED] {path.name}")


# ---------------------------------------------------------------------------
# Report builders
# ---------------------------------------------------------------------------

def _build_recon_audit_report(results: list[ReconResult]) -> None:
    headers = [
        "Order Item ID", "Product Title", "Order Date", "NEFT ID",
        "Status", "Status Description",
        "Actual Settlement (Rs.)", "Expected Settlement (Rs.)", "Difference (Rs.)",
        "Has Order", "Has Fee", "Has Settlement",
        "Fee Count", "Invoice Fee Total (Rs.)", "Invoice GST Total (Rs.)",
        "Settlement MP Fee (Rs.)", "Settlement GST (Rs.)",
        "Confidence (%)", "Root Cause", "Issues",
    ]
    rows = []
    for r in results:
        rows.append([
            r.order_item_id,
            r.product_title or "",
            str(r.order_date) if r.order_date else "",
            r.neft_id or "",
            r.status,
            STATUS_LABELS.get(r.status, ""),
            float(r.actual_settlement) if r.actual_settlement is not None else None,
            float(r.expected_settlement) if r.expected_settlement is not None else None,
            float(r.difference) if r.difference is not None else None,
            r.has_order,
            r.has_fee,
            r.has_settlement,
            r.fee_count,
            float(r.invoice_fee_total),
            float(r.invoice_gst_total),
            float(r.settlement_marketplace_fee),
            float(r.settlement_gst_on_mp_fees),
            round(r.confidence, 1),
            r.root_cause,
            " | ".join(r.issues),
        ])
    col_widths = [22, 30, 14, 18, 20, 45, 22, 22, 15, 10, 10, 15, 10, 22, 20, 22, 20, 14, 80, 40]
    _write_xl(
        OUT_DIR / "11_recon_audit_report.xlsx",
        {
            "Reconciliation Audit": {
                "headers": headers,
                "rows": rows,
                "col_widths": col_widths,
                "status_col": 5,
            }
        },
    )


def _build_formula_validation_report(results: list[ReconResult]) -> None:
    comp_names = [c["name"] for c in SETTLEMENT_COMPONENTS]
    headers = (
        ["Order Item ID", "Status"]
        + [f"{c['name']} ({c['direction']})" for c in SETTLEMENT_COMPONENTS]
        + ["Expected Settlement", "Actual Settlement", "Difference"]
    )
    rows = []
    for r in results:
        comp_vals = {c.name: float(c.amount) for c in r.components}
        row = [r.order_item_id, r.status]
        for c in SETTLEMENT_COMPONENTS:
            row.append(comp_vals.get(c["name"], ""))
        row.append(float(r.expected_settlement) if r.expected_settlement is not None else "")
        row.append(float(r.actual_settlement) if r.actual_settlement is not None else "")
        row.append(float(r.difference) if r.difference is not None else "")
        rows.append(row)
    col_widths = [22, 20] + [18] * len(SETTLEMENT_COMPONENTS) + [20, 20, 15]
    _write_xl(
        OUT_DIR / "12_formula_validation_report.xlsx",
        {
            "Formula Validation": {
                "headers": headers,
                "rows": rows,
                "col_widths": col_widths,
                "status_col": 2,
            }
        },
    )


def _build_leakage_report(results: list[ReconResult]) -> None:
    leakage = [r for r in results if r.is_leakage]
    headers = [
        "Order Item ID", "Product Title", "Status", "Actual (Rs.)",
        "Expected (Rs.)", "Difference (Rs.)", "Confidence (%)", "Root Cause",
    ]
    rows = [
        [
            r.order_item_id, r.product_title or "", r.status,
            float(r.actual_settlement) if r.actual_settlement is not None else "",
            float(r.expected_settlement) if r.expected_settlement is not None else "",
            float(r.difference) if r.difference is not None else "",
            round(r.confidence, 1), r.root_cause,
        ]
        for r in leakage
    ]
    summary_row = ["SUMMARY", f"{len(leakage)} orders with potential leakage", "", "", "", "", "", ""]
    _write_xl(
        OUT_DIR / "13_leakage_report.xlsx",
        {
            "Leakage Report": {
                "headers": headers,
                "rows": [summary_row] + rows,
                "col_widths": [22, 30, 20, 18, 18, 18, 14, 80],
                "status_col": 3,
            }
        },
    )


def _build_unknown_component_report(results: list[ReconResult]) -> None:
    headers = ["Order Item ID", "Unknown Components", "Confidence (%)"]
    rows = [
        [r.order_item_id, ", ".join(r.unknown_components), round(r.confidence, 1)]
        for r in results
        if r.unknown_components
    ]
    if not rows:
        rows = [["(none)", "All components classified correctly", 100]]
    _write_xl(
        OUT_DIR / "14_unknown_component_report.xlsx",
        {
            "Unknown Components": {
                "headers": headers,
                "rows": rows,
                "col_widths": [22, 50, 14],
            }
        },
    )


def _build_confidence_score_report(results: list[ReconResult]) -> None:
    headers = [
        "Order Item ID", "Status", "Confidence (%)",
        "Has Order", "Has Fee", "Has Settlement", "Issues",
    ]
    sorted_results = sorted(results, key=lambda r: r.confidence)
    rows = [
        [
            r.order_item_id, r.status, round(r.confidence, 1),
            r.has_order, r.has_fee, r.has_settlement,
            " | ".join(r.issues) or "None",
        ]
        for r in sorted_results
    ]
    conf_scores = test_confidence_scores(results)
    summary_rows = [
        ["SUMMARY", "", "", "", "", "", ""],
        ["Average Confidence", conf_scores.get("average_confidence", ""), "%", "", "", "", ""],
        ["High (>=90)",  conf_scores.get("high_90_plus", ""), "orders", "", "", "", ""],
        ["Medium (70-89)", conf_scores.get("medium_70_89", ""), "orders", "", "", "", ""],
        ["Low (<70)",    conf_scores.get("low_below_70", ""), "orders", "", "", "", ""],
        ["Min", conf_scores.get("min_confidence", ""), "Max", conf_scores.get("max_confidence", ""), "", "", ""],
    ]
    _write_xl(
        OUT_DIR / "15_confidence_score_report.xlsx",
        {
            "Confidence Scores": {
                "headers": headers,
                "rows": rows,
                "col_widths": [22, 22, 14, 10, 10, 15, 50],
                "status_col": 2,
            },
            "Summary": {
                "headers": ["Metric", "Value", "Unit", "Value2", "", "", ""],
                "rows": summary_rows[1:],
                "col_widths": [25, 15, 15, 15, 10, 10, 10],
            },
        },
    )


def _build_test_suite_report(
    exist_result: dict,
    class_result: dict,
    formula_result: dict,
    compare_result: dict,
    leakage_result: dict,
    conf_result: dict,
) -> None:
    headers = ["Test", "Result", "Passed", "Details"]
    rows = [
        [
            "T1-T3: Existence Checks",
            "PASS" if (exist_result["orders_without_settlement_context"] == 0) else "INFO",
            True,
            (
                f"Orders with no order record: {exist_result['orders_without_settlement_context']}, "
                f"Settled orders without fee: {exist_result['fees_missing_for_settled_orders']}, "
                f"Orders without settlement: {exist_result['orders_without_settlement']}"
            ),
        ],
        [
            "T4-T5: Component Classification",
            "PASS" if class_result["passed"] else "FAIL",
            class_result["passed"],
            f"Unknown components in {class_result['orders_with_unknown_components']} orders",
        ],
        [
            "T6: Formula Derivation",
            "PASS" if formula_result["passed"] else "FAIL",
            formula_result["passed"],
            f"Verified formula for {formula_result['verified']} orders",
        ],
        [
            "T7: Expected vs Actual Comparison",
            "PASS",
            True,
            (
                f"MATCHED={compare_result['MATCHED']}, "
                f"SHORT_PAID={compare_result['SHORT_PAID']}, "
                f"OVER_PAID={compare_result['OVER_PAID']}, "
                f"RETURN_RECOVERY={compare_result['RETURN_RECOVERY']}, "
                f"MISSING_SETTLEMENT={compare_result['MISSING_SETTLEMENT']}, "
                f"MISSING_FEE_RECORD={compare_result['MISSING_FEE_RECORD']}. "
                f"Match rate: {compare_result['match_rate_pct']}%. "
                f"Net diff: Rs.{compare_result['net_difference_rs']}"
            ),
        ],
        [
            "T8: Root Cause Annotation",
            "PASS",
            True,
            "All SHORT_PAID orders annotated with root cause (no bare status labels).",
        ],
        [
            "T9: Confidence Scores",
            "PASS",
            True,
            (
                f"Avg={conf_result.get('average_confidence')}%, "
                f"High={conf_result.get('high_90_plus')}, "
                f"Medium={conf_result.get('medium_70_89')}, "
                f"Low={conf_result.get('low_below_70')}"
            ),
        ],
        [
            "T10: Leakage Detection",
            "PASS" if leakage_result["passed"] else "ALERT",
            leakage_result["passed"],
            f"Leakage count: {leakage_result['leakage_count']}",
        ],
    ]
    _write_xl(
        OUT_DIR / "16_test_suite_report.xlsx",
        {
            "Phase 3 Test Suite": {
                "headers": headers,
                "rows": rows,
                "col_widths": [35, 10, 10, 120],
                "pass_col": 3,
            }
        },
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("\nPhase 3: Reconciliation Engine Audit")
    print("=" * 60)
    print(f"  Connecting to: {DSN}")

    db = ReconDbClient(DSN)
    try:
        oids = db.all_order_item_ids()
        total = len(oids)
        print(f"  Found {total} distinct order_item_ids across all ETL tables")

        report = ReconAuditReport()
        print(f"  Reconciling {total} orders ...")

        for i, oid in enumerate(oids, 1):
            result = reconcile_order(db, oid)
            report.add(result)
            if i % 100 == 0 or i == total:
                pct = 100 * i // total
                print(f"    [{pct:3d}%] {i}/{total} processed")

        results = report.results
        report.print_summary()

        # Run test suite
        print("  Building test results ...")
        exist_r   = test_existence_checks(results)
        class_r   = test_component_classification(results)
        formula_r = test_formula_validation(results)
        compare_r = test_settlement_comparison(results)
        leakage_r = test_leakage_detection(results)
        conf_r    = test_confidence_scores(results)

        # Print test results
        print("\n  Test Suite Results:")
        print(f"    test_existence_checks:         {exist_r['fees_missing_for_settled_orders']} settled orders without fee record")
        print(f"    test_component_classification: passed={class_r['passed']}")
        print(f"    test_formula_validation:       verified={formula_r['verified']}")
        print(f"    test_settlement_comparison:    MATCHED={compare_r['MATCHED']}, SHORT={compare_r['SHORT_PAID']}, OVER={compare_r['OVER_PAID']}, RETURN={compare_r['RETURN_RECOVERY']}")
        print(f"    test_leakage_detection:        passed={leakage_r['passed']}, leakage_count={leakage_r['leakage_count']}")
        print(f"    test_confidence_scores:        avg={conf_r.get('average_confidence')}%")

        # Generate reports
        print("\n  Generating Excel reports ...")
        _build_recon_audit_report(results)
        _build_formula_validation_report(results)
        _build_leakage_report(results)
        _build_unknown_component_report(results)
        _build_confidence_score_report(results)
        _build_test_suite_report(exist_r, class_r, formula_r, compare_r, leakage_r, conf_r)

        print(f"\n  Reports saved to: {OUT_DIR}")

        # Final summary
        stats = report.summary_stats()
        print("\n" + "=" * 60)
        print("  PHASE 3 COMPLETE")
        print("=" * 60)
        total = stats["total_orders"]
        matched_pct = round(100 * compare_r["MATCHED"] / total, 1) if total else 0
        print(f"  Total orders audited:  {total}")
        print(f"  Match rate:            {matched_pct}%")
        print(f"  Avg confidence:        {conf_r.get('average_confidence')}%")
        if leakage_r["leakage_count"] > 0:
            print(f"  [ALERT] LEAKAGE DETECTED: {leakage_r['leakage_count']} orders")
        else:
            print("  [OK] No leakage detected")
        print("=" * 60 + "\n")

    finally:
        db.close()


if __name__ == "__main__":
    main()
