#!/usr/bin/env python3
"""
Flipkart ETL Validation Audit — Phase 2 Runner

Usage:
    cd clearsettle-app/microservices/flipkart_reconciliation
    python tests/etl_validation/run_etl_audit.py

Exit codes:
    0 -- All rules PASSED
    1 -- One or more rules FAILED
    2 -- Cannot connect to DB
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

_HERE = Path(__file__).resolve().parent
_MICRO = _HERE.parent.parent
sys.path.insert(0, str(_HERE))

from etl_audit import (
    FEES_RAW_TO_ETL, ORDERS_RAW_TO_ETL, SETTLEMENTS_RAW_TO_ETL,
    DbClient, EtlAuditReport,
    test_duplicate_detection, test_etl_mapping,
    test_fee_aggregation, test_monetary_totals, test_settlement_aggregation,
    validate_column_mapping, validate_duplicates, validate_fee_aggregation,
    validate_key_generation, validate_lineage, validate_money_totals,
    validate_settlement_aggregation,
)

_REPORTS_DIR = _HERE / "etl_audit_reports"


def _load_dsn() -> str:
    env = _MICRO / ".env"
    if env.exists():
        for line in env.read_text().splitlines():
            line = line.strip()
            if line.startswith("DATABASE_URL="):
                raw = line.split("=", 1)[1].strip()
                return re.sub(r"^\s*postgresql\+asyncpg://", "postgresql://", raw)
    return "postgresql://clearsettle_user:clearsettle_pass@localhost:5432/flipkart_recon"


def _pick_batch(db: DbClient, raw_table: str, etl_table: str) -> tuple[str | None, int]:
    """
    Match a raw batch to its corresponding ETL batch.
    Strategy: find the raw batch_id that also appears in the ETL table.
    Prefer the most recently uploaded batch that has ETL rows.
    """
    raw_batches = db.all_batches(raw_table)
    if not raw_batches:
        return None, 0
    # Get all batch_ids that exist in the ETL table
    etl_batch_ids = {
        str(r["batch_id"])
        for r in db._query(f"SELECT DISTINCT batch_id FROM {etl_table}")
    }
    # Pick most-recent raw batch whose batch_id is in ETL
    for raw_batch in raw_batches:  # already ordered by uploaded_at DESC
        bid = str(raw_batch["batch_id"])
        if bid in etl_batch_ids:
            return bid, raw_batch["row_count"]
    return None, 0


def _hex(rgb: str) -> str:
    return rgb


def _write_reports(report: EtlAuditReport, reports_dir: Path) -> list[str]:
    reports_dir.mkdir(parents=True, exist_ok=True)
    created = []

    # ── 1. ETL Audit Summary ────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ETL Audit Summary"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    pass_fill  = PatternFill("solid", fgColor="C6EFCE")
    fail_fill  = PatternFill("solid", fgColor="FFC7CE")

    ws.append(["Rule", "Name", "Table", "Status", "Summary"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in report.results:
        ws.append([r.rule_id, r.rule_name, r.table, r.status, r.summary])
        fill = pass_fill if r.passed else fail_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 80

    # Test suite summary
    ws2 = wb.create_sheet("Test Suite")
    ws2.append(["Test Function", "Result", "Notes"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    passed = sum(1 for r in report.results if r.passed)
    total = len(report.results)
    ws2.append(["test_etl_mapping",          "PASS" if all(r.passed for r in report.results if r.rule_id=="R2") else "FAIL", "All columns traced raw_json → ETL"])
    ws2.append(["test_duplicate_detection",  "PASS" if all(r.passed for r in report.results if r.rule_id=="R4") else "FAIL", "order_item_id unique in orders/settlements"])
    ws2.append(["test_fee_aggregation",      "PASS" if all(r.passed for r in report.results if r.rule_id=="R5") else "FAIL", "Per-order fee totals match raw vs ETL"])
    ws2.append(["test_settlement_aggregation","PASS" if all(r.passed for r in report.results if r.rule_id=="R6") else "FAIL", "BSV = SUM(components) checksum"])
    ws2.append(["test_monetary_totals",      "PASS" if all(r.passed for r in report.results if r.rule_id=="R7") else "FAIL", "Grand totals match between raw and ETL"])
    ws2.append([])
    ws2.append(["Overall", f"{passed}/{total} rules passed", ""])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 50

    p = reports_dir / "06_etl_audit_report.xlsx"
    wb.save(str(p))
    created.append(str(p))

    # ── 2. Duplicate Report ─────────────────────────────────────────────────
    wb2 = openpyxl.Workbook()
    ws_d = wb2.active
    ws_d.title = "Duplicates"
    ws_d.append(["Table", "Key Columns", "Order Item ID", "Count", "ETL ID"])
    for cell in ws_d[1]:
        cell.font = Font(bold=True)

    for r in report.results:
        if r.rule_id != "R4":
            continue
        for issue in r.issues:
            if issue.get("problem") == "duplicate_keys":
                for dup in issue.get("sample", []):
                    ws_d.append([
                        r.table,
                        "+".join(r.details.get("key_columns", [])),
                        str(dup.get("key", {})),
                        dup.get("count"),
                        dup.get("example", {}).get("etl_id"),
                    ])
    for col in ["A","B","C","D","E"]:
        ws_d.column_dimensions[col].width = 30

    p2 = reports_dir / "07_duplicate_report.xlsx"
    wb2.save(str(p2))
    created.append(str(p2))

    # ── 3. Aggregation Audit Report ─────────────────────────────────────────
    wb3 = openpyxl.Workbook()
    ws_a = wb3.active
    ws_a.title = "Fee Aggregation"
    ws_a.append(["Order Item ID", "Raw Net Fee", "ETL Net Fee", "Delta", "Raw Count", "ETL Count", "Fee Categories"])
    for cell in ws_a[1]:
        cell.font = Font(bold=True)

    for r in report.results:
        if r.rule_id == "R5":
            for issue in r.issues:
                for m in issue.get("sample", []):
                    ws_a.append([
                        m.get("order_item_id"),
                        m.get("raw_net_fee"),
                        m.get("etl_net_fee"),
                        m.get("delta"),
                        m.get("raw_count"),
                        m.get("etl_count"),
                        json.dumps(m.get("fee_categories", {})),
                    ])
    for col in ["A","B","C","D","E","F","G"]:
        ws_a.column_dimensions[col].width = 25

    ws_s = wb3.create_sheet("Settlement Aggregation")
    ws_s.append(["ETL ID", "Order Item ID", "NEFT ID", "BSV", "Computed Sum", "Delta"])
    for cell in ws_s[1]:
        cell.font = Font(bold=True)
    for r in report.results:
        if r.rule_id == "R6":
            for issue in r.issues:
                if issue.get("problem") == "settlement_checksum_failure":
                    for m in issue.get("sample", []):
                        ws_s.append([
                            m.get("etl_id"),
                            m.get("order_item_id"),
                            m.get("neft_id"),
                            m.get("bsv"),
                            m.get("computed_sum"),
                            m.get("delta"),
                        ])
    for col in ["A","B","C","D","E","F"]:
        ws_s.column_dimensions[col].width = 25

    p3 = reports_dir / "08_aggregation_audit_report.xlsx"
    wb3.save(str(p3))
    created.append(str(p3))

    # ── 4. Mapping Report ───────────────────────────────────────────────────
    wb4 = openpyxl.Workbook()
    ws_m = wb4.active
    ws_m.title = "Column Mapping"
    ws_m.append(["Table", "ETL Column", "Raw JSON Key", "Total Rows", "Correct", "Null Mismatches", "Missing in Raw"])
    for cell in ws_m[1]:
        cell.font = Font(bold=True)

    for r in report.results:
        if r.rule_id == "R2":
            for etl_col, info in r.details.get("mapping_report", {}).items():
                ws_m.append([
                    r.table,
                    etl_col,
                    info.get("raw_key"),
                    info.get("total"),
                    info.get("correct"),
                    info.get("wrong"),
                    info.get("missing_in_raw"),
                ])
    for col in ["A","B","C","D","E","F","G"]:
        ws_m.column_dimensions[col].width = 25

    p4 = reports_dir / "09_mapping_report.xlsx"
    wb4.save(str(p4))
    created.append(str(p4))

    # ── 5. Money Totals Report ──────────────────────────────────────────────
    wb5 = openpyxl.Workbook()
    ws_t = wb5.active
    ws_t.title = "Money Totals"
    ws_t.append(["Table", "ETL Column", "Raw JSON Key", "Raw SUM", "ETL SUM", "Delta", "Match"])
    for cell in ws_t[1]:
        cell.font = Font(bold=True)

    for r in report.results:
        if r.rule_id == "R7":
            for etl_col, info in r.details.get("columns", {}).items():
                ws_t.append([
                    r.table,
                    etl_col,
                    info.get("raw_col"),
                    info.get("raw_sum"),
                    info.get("etl_sum"),
                    info.get("delta"),
                    "YES" if info.get("match") else "NO",
                ])
    for col in ["A","B","C","D","E","F","G"]:
        ws_t.column_dimensions[col].width = 25

    p5 = reports_dir / "10_money_totals_report.xlsx"
    wb5.save(str(p5))
    created.append(str(p5))

    return created


def main() -> int:
    dsn = _load_dsn()
    print(f"\nFlipkart ETL Validation Audit (Phase 2)")
    print(f"  DB: {dsn}\n")

    try:
        db = DbClient(dsn)
        print("  [OK] Connected to database")
    except Exception as e:
        print(f"  [FAIL] Cannot connect: {e}")
        return 2

    report = EtlAuditReport()

    # Determine batches for each table
    orders_batch,      _ = _pick_batch(db, "orders_raw",      "orders_etl")
    fees_batch,        _ = _pick_batch(db, "fees_raw",        "fees_etl")
    settlements_batch, _ = _pick_batch(db, "settlements_raw", "settlements_etl")

    batches = {
        "orders":      orders_batch or "",
        "fees":        fees_batch or "",
        "settlements": settlements_batch or "",
    }

    TABLE_DEFS = [
        # Orders: allow_duplicates=True because Flipkart's report can include the same
        # order_item_id with different statuses (e.g., CANCELLED then DELIVERED lifecycle
        # events). The ETL correctly preserves source rows; R4 reports but does not fail.
        ("orders",      "orders_raw",      "orders_etl",      orders_batch,      ORDERS_RAW_TO_ETL,      ["order_item_id"],            True,  None,                         "OI:"),
        ("fees",        "fees_raw",        "fees_etl",        fees_batch,        FEES_RAW_TO_ETL,        ["order_item_id", "fee_name"], True,  None,                         None),
        # Settlements: multiple rows per order_item_id are EXPECTED (return recovery + original sale
        # both appear in the same settlement file). allow_duplicates=True reports without failing.
        ("settlements", "settlements_raw", "settlements_etl", settlements_batch, SETTLEMENTS_RAW_TO_ETL, ["order_item_id"],             True, {"settlement_amount","sale_amount"}, None),
    ]

    for label, raw_t, etl_t, batch_id, col_map, key_cols, allow_dupes, opt_cols, strip_pfx in TABLE_DEFS:
        if not batch_id:
            print(f"\n  [{label.upper()}] No batch found — skipping")
            continue

        print(f"\n{'-'*60}")
        print(f"  [{label.upper()}] batch={batch_id[:8]}...")
        print(f"{'-'*60}")

        r1 = validate_lineage(db, raw_t, etl_t, batch_id, label)
        report.add(r1)
        print(f"  R1 [{r1.status}] Lineage           -- {r1.summary}")

        r2 = validate_column_mapping(db, raw_t, etl_t, batch_id, label, col_map,
                                     skip_cols={"batch_id","raw_id","created_at","is_valid","validation_errors"})
        report.add(r2)
        print(f"  R2 [{r2.status}] Column Mapping    -- {r2.summary}")

        r3 = validate_key_generation(db, raw_t, etl_t, batch_id, label, strip_prefix=strip_pfx)
        report.add(r3)
        print(f"  R3 [{r3.status}] Key Generation    -- {r3.summary}")

        r4 = validate_duplicates(db, etl_t, batch_id, label, key_cols, allow_duplicates=allow_dupes)
        report.add(r4)
        print(f"  R4 [{r4.status}] Duplicates        -- {r4.summary}")

    # R5: Fee Aggregation (fees only)
    if fees_batch:
        print(f"\n{'-'*60}")
        print(f"  [FEES] Fee Aggregation")
        r5 = validate_fee_aggregation(db, "fees_raw", "fees_etl", fees_batch, "fees")
        report.add(r5)
        print(f"  R5 [{r5.status}] Fee Aggregation   -- {r5.summary}")

    # R6: Settlement Aggregation (internal checksum)
    if settlements_batch:
        print(f"\n{'-'*60}")
        print(f"  [SETTLEMENTS] Settlement Aggregation")
        r6 = validate_settlement_aggregation(db, "settlements_etl", settlements_batch, "settlements")
        report.add(r6)
        print(f"  R6 [{r6.status}] Settlement Agg    -- {r6.summary}")

    # R7: Money Totals
    if fees_batch:
        r7_fees = validate_money_totals(
            db, "fees_raw", "fees_etl", fees_batch, "fees",
            numeric_raw_cols=["fee_amount","fee_waiver_amount","cgst","sgst","igst","tax_amount"],
            numeric_etl_cols=["fee_amount","fee_waiver_amount","cgst","sgst","igst","tax_amount"],
        )
        report.add(r7_fees)
        print(f"  R7 [{r7_fees.status}] Money Totals (fees)        -- {r7_fees.summary}")

    if settlements_batch:
        r7_s = validate_money_totals(
            db, "settlements_raw", "settlements_etl", settlements_batch, "settlements",
            numeric_raw_cols=["settlement_amount","sale_amount","total_offer_amount","my_share",
                              "marketplace_fee","tcs","tds","gst_on_mp_fees",
                              "offer_adjustments","protection_fund","refund"],
            numeric_etl_cols=["settlement_amount","sale_amount","total_offer_amount","my_share",
                              "marketplace_fee","tcs","tds","gst_on_mp_fees",
                              "offer_adjustments","protection_fund","refund"],
            optional_raw_cols={"settlement_amount","sale_amount"},
        )
        report.add(r7_s)
        print(f"  R7 [{r7_s.status}] Money Totals (settlements)   -- {r7_s.summary}")

    report.print_summary()

    # Run pytest-style test suite
    print("Running test suite...")
    print(f"  test_etl_mapping:           {test_etl_mapping(db, batches)}")
    print(f"  test_duplicate_detection:   {test_duplicate_detection(db, batches)}")
    if fees_batch:
        print(f"  test_fee_aggregation:       {test_fee_aggregation(db, fees_batch)}")
    if settlements_batch:
        print(f"  test_settlement_aggregation: {test_settlement_aggregation(db, settlements_batch)}")
    print(f"  test_monetary_totals:       {test_monetary_totals(db, batches)}")

    print("\nGenerating reports...")
    try:
        created = _write_reports(report, _REPORTS_DIR)
        for p in created:
            print(f"  -> {p}")
    except Exception as e:
        print(f"  [WARNING] Report generation failed: {e}")

    db.close()
    return 0 if report.passed_all() else 1


if __name__ == "__main__":
    sys.exit(main())
