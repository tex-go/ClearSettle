"""
File Fingerprinting Engine.

For every uploaded file, extracts:
  - sheet names (Excel) or table name (CSV)
  - raw header rows (first 20 rows per sheet — wider window for P&L-style files)
  - flattened column names (best-row detection instead of first-row assumption)
  - sample data rows (first 20)
  - content tokens — all meaningful cell values from first 100 rows per sheet
  - file metadata

Produces a FileFingerprint dataclass — the input contract for
platform_detector, report_type_detector, and schema_detector.

Key design: content_tokens captures cell VALUES (not just column headers), which
is critical for Flipkart P&L reports where headers like "Profit & Loss Report",
"Earnings on Platform" and "PNL Summary" are row-values, not column names.
"""
from __future__ import annotations

import csv
import hashlib
import io
import logging
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

_EXCEL_EXTS = frozenset([".xlsx", ".xls", ".xlsm", ".xlsb"])
_CSV_EXTS   = frozenset([".csv", ".tsv", ".txt"])

_MAX_HEADER_ROWS      = 20   # scan 20 rows to find the actual header row
_MAX_SAMPLE_ROWS      = 20   # collect 20 data rows as samples
_MAX_CONTENT_SCAN_ROWS = 100  # scan first 100 rows per sheet for content tokens


@dataclass
class SheetFingerprint:
    sheet_name:     str
    raw_headers:    List[List[str]]        # up to first _MAX_HEADER_ROWS rows
    flat_columns:   List[str]              # best-row-detected column names
    sample_rows:    List[Dict[str, Any]]   # first N data rows as dicts
    row_count:      int
    col_count:      int
    content_tokens: List[str] = field(default_factory=list)  # cell values from first 100 rows


@dataclass
class FileFingerprint:
    file_name:          str
    file_hash_sha256:   str
    mime_type:          str
    file_size_bytes:    int
    sheets:             List[SheetFingerprint]
    all_column_names:   List[str]            # deduplicated union of flat_columns across sheets
    schema_signature:   str                  # SHA-256 of sorted, cleaned all_column_names
    is_excel:           bool
    is_csv:             bool
    all_content_tokens: List[str] = field(default_factory=list)  # union of all sheets' tokens
    parse_warnings:     List[str] = field(default_factory=list)


def fingerprint_file(
    file_bytes: bytes,
    file_name:  str,
) -> FileFingerprint:
    """
    Produce a FileFingerprint from raw file bytes.
    Never raises — parse failures are captured in parse_warnings.
    """
    sha256   = hashlib.sha256(file_bytes).hexdigest()
    ext      = _get_ext(file_name)
    is_excel = ext in _EXCEL_EXTS
    is_csv   = ext in _CSV_EXTS
    mime     = _mime_from_ext(ext)
    warnings: List[str] = []
    sheets: List[SheetFingerprint] = []

    if is_excel:
        sheets, warnings = _fingerprint_excel(file_bytes, warnings)
    elif is_csv:
        sheets, warnings = _fingerprint_csv(file_bytes, file_name, warnings)
    else:
        warnings.append(f"Unsupported file extension '{ext}'; attempting Excel parse.")
        sheets, warnings = _fingerprint_excel(file_bytes, warnings)

    all_cols   = _collect_all_columns(sheets)
    all_tokens = _collect_all_tokens(sheets)
    sig        = _schema_signature(all_cols)

    return FileFingerprint(
        file_name=file_name,
        file_hash_sha256=sha256,
        mime_type=mime,
        file_size_bytes=len(file_bytes),
        sheets=sheets,
        all_column_names=all_cols,
        schema_signature=sig,
        is_excel=is_excel,
        is_csv=is_csv,
        all_content_tokens=all_tokens,
        parse_warnings=warnings,
    )


# ── Internal helpers ──────────────────────────────────────────────────────────

def _get_ext(name: str) -> str:
    name = name.lower().strip()
    if "." not in name:
        return ""
    return "." + name.rsplit(".", 1)[-1]


def _mime_from_ext(ext: str) -> str:
    mapping = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls":  "application/vnd.ms-excel",
        ".csv":  "text/csv",
        ".tsv":  "text/tab-separated-values",
        ".txt":  "text/plain",
    }
    return mapping.get(ext, "application/octet-stream")


def _clean_cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip().replace("\n", " ").replace("\r", "")


def _is_content_token(v: str) -> bool:
    """Return True if a cell value is a meaningful content token worth indexing."""
    v = v.strip().lower()
    if len(v) < 3:
        return False
    # Skip purely numeric values (dates formatted as numbers, amounts, etc.)
    cleaned = v.replace(",", "").replace(".", "").replace("-", "").replace("/", "").replace("%", "").replace("₹", "").replace(" ", "")
    if cleaned.isnumeric():
        return False
    return True


def _fingerprint_excel(
    file_bytes: bytes,
    warnings: List[str],
) -> tuple[List[SheetFingerprint], List[str]]:
    try:
        import openpyxl
    except ImportError:
        warnings.append("openpyxl not installed; cannot fingerprint Excel.")
        return [], warnings

    sheets: List[SheetFingerprint] = []
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        warnings.append(f"openpyxl failed to open workbook: {exc}")
        return [], warnings

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)

            raw_header_rows: List[List[str]] = []
            data_rows: List[List[Any]] = []
            content_token_set: set = set()
            total = 0

            for row in rows_iter:
                cleaned = [_clean_cell(c) for c in row]
                total += 1

                if total <= _MAX_HEADER_ROWS:
                    raw_header_rows.append(cleaned)
                elif total <= _MAX_HEADER_ROWS + _MAX_SAMPLE_ROWS:
                    data_rows.append(list(row))

                # Collect content tokens from ALL cells in first 100 rows
                if total <= _MAX_CONTENT_SCAN_ROWS:
                    for cell_val in cleaned:
                        if _is_content_token(cell_val):
                            content_token_set.add(cell_val.strip().lower())

            flat_cols = _flatten_headers(raw_header_rows)
            sample    = _rows_to_dicts(flat_cols, data_rows)
            col_count = len(flat_cols) if flat_cols else (len(raw_header_rows[0]) if raw_header_rows else 0)

            sheets.append(SheetFingerprint(
                sheet_name=sheet_name,
                raw_headers=raw_header_rows,
                flat_columns=flat_cols,
                sample_rows=sample,
                row_count=total,
                col_count=col_count,
                content_tokens=sorted(content_token_set),
            ))
        except Exception as exc:
            warnings.append(f"Sheet '{sheet_name}' parse error: {exc}")

    return sheets, warnings


def _fingerprint_csv(
    file_bytes: bytes,
    file_name: str,
    warnings: List[str],
) -> tuple[List[SheetFingerprint], List[str]]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        warnings.append("Could not decode CSV with any known encoding.")
        return [], warnings

    delimiter = "\t" if file_name.lower().endswith(".tsv") else _detect_delimiter(text)

    try:
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        all_rows = list(reader)
    except Exception as exc:
        warnings.append(f"CSV parse error: {exc}")
        return [], warnings

    if not all_rows:
        return [], warnings

    header_rows   = [[ _clean_cell(c) for c in r ] for r in all_rows[:_MAX_HEADER_ROWS]]
    data_rows_raw = all_rows[_MAX_HEADER_ROWS: _MAX_HEADER_ROWS + _MAX_SAMPLE_ROWS]

    # Collect content tokens from first 100 rows
    content_token_set: set = set()
    for row in all_rows[:_MAX_CONTENT_SCAN_ROWS]:
        for cell_val in row:
            v = _clean_cell(cell_val)
            if _is_content_token(v):
                content_token_set.add(v.strip().lower())

    flat_cols = _flatten_headers(header_rows)
    sample    = _rows_to_dicts(flat_cols, [r for r in data_rows_raw])
    col_count = len(flat_cols) or (len(all_rows[0]) if all_rows else 0)

    sheet = SheetFingerprint(
        sheet_name="sheet1",
        raw_headers=header_rows,
        flat_columns=flat_cols,
        sample_rows=sample,
        row_count=len(all_rows),
        col_count=col_count,
        content_tokens=sorted(content_token_set),
    )
    return [sheet], warnings


def _detect_delimiter(text: str) -> str:
    sample = text[:4096]
    tab_count   = sample.count("\t")
    comma_count = sample.count(",")
    pipe_count  = sample.count("|")
    if tab_count > comma_count and tab_count > pipe_count:
        return "\t"
    if pipe_count > comma_count:
        return "|"
    return ","


def _flatten_headers(raw_rows: List[List[str]]) -> List[str]:
    """
    Find the best header row in raw_rows and return flattened column names.

    Strategy:
    1. Find the row with the MOST non-empty cells — this is the likely header row.
       This handles P&L-style reports where metadata rows precede the actual column headers.
    2. Check if the row immediately before is a section-header row (Flipkart payment
       report style: section row + column name row). If so, combine them.
    3. Fall back to the first non-empty row if no clear winner.
    """
    if not raw_rows:
        return []

    # Find the row with the most non-empty cells
    best_idx   = 0
    best_count = 0
    for i, row in enumerate(raw_rows):
        count = sum(1 for c in row if c.strip())
        if count > best_count:
            best_count = count
            best_idx   = i

    if best_count == 0:
        return []

    if best_count == 1:
        # Only single-cell rows — return whatever non-empty values exist
        for row in raw_rows:
            if any(c.strip() for c in row):
                return [_norm_col(c) for c in row if c.strip()]
        return []

    best_row = raw_rows[best_idx]

    # Check for Flipkart-style 2-row merged header:
    # prev row has ≥2 cells and its non-empty count is proportional to best row
    if best_idx > 0:
        prev_row   = raw_rows[best_idx - 1]
        prev_count = sum(1 for c in prev_row if c.strip())
        if prev_count >= 2:
            # Build combined section+column names with forward-fill
            section_row = list(prev_row)
            col_row     = list(best_row)
            # Pad to same length
            n = max(len(section_row), len(col_row))
            section_row += [""] * (n - len(section_row))
            col_row     += [""] * (n - len(col_row))

            # Forward-fill blanks in section_row
            last = ""
            for i, v in enumerate(section_row):
                if v.strip():
                    last = v.strip()
                else:
                    section_row[i] = last

            result: List[str] = []
            seen:   Dict[str, int] = {}
            for i in range(n):
                sec = section_row[i].strip()
                col = col_row[i].strip()
                if not sec and not col:
                    continue
                combined = _norm_col(f"{sec} {col}".strip() if sec and col else sec or col)
                if combined in seen:
                    seen[combined] += 1
                    combined = f"{combined}_{seen[combined]}"
                else:
                    seen[combined] = 0
                result.append(combined)
            if result:
                return result

    # Single-row header: use the best row directly
    result = []
    seen: Dict[str, int] = {}
    for c in best_row:
        if not c.strip():
            continue
        norm = _norm_col(c)
        if norm in seen:
            seen[norm] += 1
            norm = f"{norm}_{seen[norm]}"
        else:
            seen[norm] = 0
        result.append(norm)
    return result


def _norm_col(name: str) -> str:
    """Lowercase, strip, collapse spaces."""
    return " ".join(name.lower().split())


def _rows_to_dicts(cols: List[str], data_rows: List[List[Any]]) -> List[Dict[str, Any]]:
    result = []
    for row in data_rows:
        d: Dict[str, Any] = {}
        for i, col in enumerate(cols):
            d[col] = row[i] if i < len(row) else None
        result.append(d)
    return result


def _collect_all_columns(sheets: List[SheetFingerprint]) -> List[str]:
    seen: set = set()
    result: List[str] = []
    for sh in sheets:
        for col in sh.flat_columns:
            if col not in seen:
                seen.add(col)
                result.append(col)
    return result


def _collect_all_tokens(sheets: List[SheetFingerprint]) -> List[str]:
    """Union of all content_tokens across all sheets (deduplicated)."""
    seen: set = set()
    result: List[str] = []
    for sh in sheets:
        for tok in sh.content_tokens:
            if tok not in seen:
                seen.add(tok)
                result.append(tok)
    return result


def _schema_signature(columns: List[str]) -> str:
    canonical = "|".join(sorted(columns))
    return hashlib.sha256(canonical.encode()).hexdigest()
